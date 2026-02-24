#!/usr/bin/env python3
"""
inventory-excel.py
==================
Parse the two S2i Excel files and produce a consolidated JSON inventory.

Inputs:
  1) S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx  (3 sheets)
  2) S2i_KBOP DB 보유 국제대회 목록_260220.xlsx          (1 sheet)

Output:
  raw/excel-inventory.json
"""

import json
import os
import sys
from pathlib import Path
import openpyxl


BASE_DIR = str(Path(__file__).resolve().parent.parent)
FILE_COMPARISON = os.path.join(
    BASE_DIR, "references",
    "S2i_KBOP \ubc31\uc5c5 DB \ud14c\uc774\ube14\ubcc4 \ub370\uc774\ud130 \ud655\uc778_260220.xlsx",
)
FILE_INTL = os.path.join(
    BASE_DIR, "references",
    "S2i_KBOP DB \ubcf4\uc720 \uad6d\uc81c\ub300\ud68c \ubaa9\ub85d_260220.xlsx",
)
OUTPUT_JSON = os.path.join(BASE_DIR, "raw", "excel-inventory.json")


def safe_int(val):
    """Convert a value to int if possible, else return None."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def safe_str(val):
    """Convert a value to stripped string, or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def safe_bool(val):
    """Convert True/False to bool, or None."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("true", "1", "o", "yes"):
        return True
    if s in ("false", "0", "x", "no"):
        return False
    return None


def resolve_merged_value(ws, row_idx, col_idx):
    """Resolve value from merged cells by finding the top-left cell."""
    val = ws.cell(row=row_idx, column=col_idx).value
    if val is not None:
        return val
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row_idx <= merged_range.max_row and
                merged_range.min_col <= col_idx <= merged_range.max_col):
            return ws.cell(
                row=merged_range.min_row, column=merged_range.min_col
            ).value
    return None


def parse_comparison_file(filepath):
    """Parse the backup-vs-S2i comparison workbook (3 sheets)."""
    print("[1/2] Parsing: {}".format(os.path.basename(filepath)))
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_data = {}
    sheet_counts = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []

        # Row 1-2: headers; data starts from row 3
        for row_idx in range(3, ws.max_row + 1):
            backup_db_name      = safe_str(resolve_merged_value(ws, row_idx, 1))   # A
            backup_no           = safe_int(ws.cell(row=row_idx, column=2).value)   # B
            backup_table_name   = safe_str(ws.cell(row=row_idx, column=3).value)   # C
            backup_column_count = safe_int(ws.cell(row=row_idx, column=4).value)   # D
            backup_row_count    = safe_int(ws.cell(row=row_idx, column=5).value)   # E
            s2i_db_name         = safe_str(ws.cell(row=row_idx, column=6).value)   # F
            s2i_table_name      = safe_str(ws.cell(row=row_idx, column=7).value)   # G
            s2i_column_count    = safe_int(ws.cell(row=row_idx, column=8).value)   # H
            s2i_row_count       = safe_int(ws.cell(row=row_idx, column=9).value)   # I
            column_match        = safe_bool(ws.cell(row=row_idx, column=10).value) # J
            row_match           = safe_bool(ws.cell(row=row_idx, column=11).value) # K
            note                = safe_str(ws.cell(row=row_idx, column=12).value)  # L
            reason              = safe_str(ws.cell(row=row_idx, column=13).value)  # M

            # Skip empty rows
            if backup_table_name is None and backup_no is None:
                continue

            rows.append({
                "backup_db_name": backup_db_name,
                "backup_no": backup_no,
                "backup_table_name": backup_table_name,
                "backup_column_count": backup_column_count,
                "backup_row_count": backup_row_count,
                "s2i_db_name": s2i_db_name,
                "s2i_table_name": s2i_table_name,
                "s2i_column_count": s2i_column_count,
                "s2i_row_count": s2i_row_count,
                "column_match": column_match,
                "row_match": row_match,
                "note": note,
                "reason": reason,
            })

        sheets_data[sheet_name] = rows
        sheet_counts[sheet_name] = len(rows)
        print("       Sheet [{}]: {} rows".format(sheet_name, len(rows)))

    total = sum(sheet_counts.values())
    print("       Total: {} rows across {} sheets".format(total, len(wb.sheetnames)))
    wb.close()
    return {"sheets": sheets_data, "summary": {"total_rows": total, "by_sheet": sheet_counts}}


def parse_international_file(filepath):
    """Parse the international tournaments workbook."""
    print("\n[2/2] Parsing: {}".format(os.path.basename(filepath)))
    wb = openpyxl.load_workbook(filepath, data_only=True)
    tournaments = []
    ws = wb[wb.sheetnames[0]]

    # Row 1: empty, Row 2: headers, Data from row 3
    for row_idx in range(3, ws.max_row + 1):
        year = ws.cell(row=row_idx, column=2).value  # B
        name = safe_str(ws.cell(row=row_idx, column=3).value)  # C
        if name is None:
            continue
        tournaments.append({"year": safe_int(year), "name": name})

    print("       Sheet [{}]: {} tournaments".format(wb.sheetnames[0], len(tournaments)))
    wb.close()
    return {"tournaments": tournaments, "count": len(tournaments)}


def print_summary(comparison_data, international_data):
    """Print a human-readable summary of all parsed data."""
    sep = "=" * 70
    print("\n" + sep)
    print("SUMMARY")
    print(sep)

    ct = comparison_data
    print("\n[\ubc31\uc5c5 DB \ud14c\uc774\ube14\ubcc4 \ub370\uc774\ud130 \ud655\uc778]")
    print("  Total rows: {}".format(ct["summary"]["total_rows"]))
    for sheet, count in ct["summary"]["by_sheet"].items():
        print("    Sheet [{}]: {} rows".format(sheet, count))

    # Per-sheet DB breakdown
    for sheet_name, rows in ct["sheets"].items():
        db_names = sorted(set(r["backup_db_name"] for r in rows if r["backup_db_name"]))
        print("\n  Sheet [{}] - Backup DBs:".format(sheet_name))
        for db in db_names:
            n = sum(1 for r in rows if r["backup_db_name"] == db)
            print("    {}: {} tables".format(db, n))

    # Match statistics
    print("\n  Column/Row match statistics:")
    for sheet_name, rows in ct["sheets"].items():
        def cnt(key, val):
            return sum(1 for r in rows if r[key] is val)
        msg = "    [{}]: Column(T={}/F={}/N={})  Row(T={}/F={}/N={})"
        print(msg.format(
            sheet_name,
            cnt("column_match", True), cnt("column_match", False), cnt("column_match", None),
            cnt("row_match", True), cnt("row_match", False), cnt("row_match", None),
        ))

    # Unique table names
    all_backup = set()
    all_s2i = set()
    for rows in ct["sheets"].values():
        for r in rows:
            if r["backup_table_name"]:
                all_backup.add(r["backup_table_name"])
            if r["s2i_table_name"]:
                all_s2i.add(r["s2i_table_name"])
    print("\n  Unique backup table names: {}".format(len(all_backup)))
    print("  Unique S2i table names: {}".format(len(all_s2i)))

    # Tables without S2i counterpart
    no_s2i = []
    for sn, rows in ct["sheets"].items():
        for r in rows:
            if r["s2i_table_name"] is None and r["backup_table_name"]:
                no_s2i.append((sn, r["backup_db_name"], r["backup_table_name"]))
    if no_s2i:
        print("\n  Tables without S2i counterpart: {}".format(len(no_s2i)))
        for sn, db, tbl in no_s2i[:10]:
            print("    [{}] {} / {}".format(sn, db, tbl))
        if len(no_s2i) > 10:
            print("    ... and {} more".format(len(no_s2i) - 10))

    # International tournaments
    intl = international_data
    print("\n[\uad6d\uc81c\ub300\ud68c \ubaa9\ub85d]")
    print("  Total tournaments: {}".format(intl["count"]))
    years = [t["year"] for t in intl["tournaments"] if t["year"]]
    if years:
        print("  Year range: {} ~ {}".format(min(years), max(years)))

    year_counts = {}
    for t in intl["tournaments"]:
        y = t["year"]
        year_counts[y] = year_counts.get(y, 0) + 1
    print("  Tournaments by year:")
    for y in sorted(year_counts.keys()):
        names = [t["name"] for t in intl["tournaments"] if t["year"] == y]
        print("    {}: {} - {}".format(y, year_counts[y], ", ".join(names)))

    print("\nDone.")


def main():
    print("=" * 70)
    print("Excel Inventory Parser")
    print("=" * 70)

    for f in [FILE_COMPARISON, FILE_INTL]:
        if not os.path.isfile(f):
            print("ERROR: File not found: {}".format(f), file=sys.stderr)
            sys.exit(1)

    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)

    comparison_data = parse_comparison_file(FILE_COMPARISON)
    international_data = parse_international_file(FILE_INTL)

    result = {
        "comparison_table": comparison_data,
        "international": international_data,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print("\nOutput written to: {}".format(OUTPUT_JSON))

    print_summary(comparison_data, international_data)


if __name__ == "__main__":
    main()
