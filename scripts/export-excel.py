#!/usr/bin/env python3
"""
데이터 사전 Excel 출력
  - column-metadata.json + upgrade-dictionary.py 매핑 기반
  - 시트 구성:
    1) 목차        : 전체 테이블 요약
    2) 컬럼 사전    : 전체 컬럼 상세 (필터용 flat 테이블)
    3) 코드값       : 주요 코드 컬럼의 허용값 목록
  - 출력: exports/데이터사전.xlsx
"""
import json, os, sys, re
from datetime import date

# openpyxl 임포트
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    sys.exit(1)

from config import (DOMAIN_TABLES, DOMAIN_LABELS, TABLE_TO_DOMAIN,
                    STANDARD_MAP, suggest_standard_name, COLUMN_DESC, BASE_DIR)

META_PATH = os.path.join(BASE_DIR, "raw", "column-metadata.json")
OUT_DIR = os.path.join(BASE_DIR, "exports")
OUT_PATH = os.path.join(OUT_DIR, "데이터사전.xlsx")
TODAY = date.today().strftime("%Y-%m-%d")

# ── 스타일 ──
HEADER_FONT = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="맑은 고딕", size=9)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
CELL_ALIGN_CENTER = Alignment(horizontal="center", vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
DOMAIN_FILLS = {
    "경기 기록": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
    "통계":     PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "실시간":   PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "마스터":   PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
}


def style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_cell(ws, row, col, center=False):
    cell = ws.cell(row=row, column=col)
    cell.font = CELL_FONT
    cell.alignment = CELL_ALIGN_CENTER if center else CELL_ALIGN
    cell.border = THIN_BORDER
    return cell


def main():
    # 데이터 로드
    with open(META_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    tables = data["tables"]

    wb = Workbook()

    # ━━━━━━━━━━━━ 시트 1: 목차 ━━━━━━━━━━━━
    ws1 = wb.active
    ws1.title = "목차"
    ws1.sheet_properties.tabColor = "2F5496"

    headers1 = ["#", "도메인", "테이블명", "표준명(안)", "컬럼 수", "행 수",
                 "PK", "대표 DB"]
    for ci, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=ci, value=h)
    style_header(ws1, 1, len(headers1))

    # 도메인 순서대로 정렬
    ordered_tables = []
    for domain in ["game", "stats", "realtime", "master"]:
        for tname in DOMAIN_TABLES.get(domain, []):
            if tname in tables:
                ordered_tables.append((domain, tname, tables[tname]))

    for ri, (domain, tname, tdata) in enumerate(ordered_tables, 2):
        domain_kr = DOMAIN_LABELS.get(domain, domain)
        std_table = tname.lower()
        vals = [
            ri - 1,
            domain_kr,
            tname,
            std_table,
            tdata["column_count"],
            f'{tdata["row_count"]:,}',
            ", ".join(tdata["pk_columns"]),
            tdata["representative_db"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = style_cell(ws1, ri, ci, center=(ci in [1, 5]))
            cell.value = v
            # 도메인 색상
            if ci == 2 and domain_kr in DOMAIN_FILLS:
                cell.fill = DOMAIN_FILLS[domain_kr]

    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 30
    ws1.column_dimensions["D"].width = 30
    ws1.column_dimensions["E"].width = 8
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 30
    ws1.column_dimensions["H"].width = 10
    ws1.column_dimensions["I"].width = 28
    ws1.auto_filter.ref = f"A1:I{len(ordered_tables)+1}"
    ws1.freeze_panes = "A2"

    # ━━━━━━━━━━━━ 시트 2: 컬럼 사전 ━━━━━━━━━━━━
    ws2 = wb.create_sheet("컬럼 사전")
    ws2.sheet_properties.tabColor = "548235"

    headers2 = ["도메인", "테이블명", "순번", "컬럼명(현행)", "표준명(안)",
                 "데이터 타입", "길이", "NULL 허용", "PK", "설명", "샘플값"]
    for ci, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=ci, value=h)
    style_header(ws2, 1, len(headers2))

    row = 2
    for domain, tname, tdata in ordered_tables:
        domain_kr = DOMAIN_LABELS.get(domain, domain)
        for col in tdata["columns"]:
            cname = col["name"]
            std_name = suggest_standard_name(cname)
            desc = COLUMN_DESC.get(cname, "")
            samples = col.get("sample_values") or []
            sample_str = ", ".join(str(s) for s in samples[:3]) if samples else ""
            length = col.get("max_length")
            length_str = str(length) if length and length > 0 else ""

            vals = [
                domain_kr,
                tname,
                col["ordinal"],
                cname,
                std_name,
                col["data_type"],
                length_str,
                "Y" if col["is_nullable"] else "N",
                "PK" if col["is_pk"] else "",
                desc,
                sample_str,
            ]
            for ci, v in enumerate(vals, 1):
                center = ci in [3, 7, 8, 9]
                cell = style_cell(ws2, row, ci, center=center)
                cell.value = v
                if ci == 1 and domain_kr in DOMAIN_FILLS:
                    cell.fill = DOMAIN_FILLS[domain_kr]
            row += 1

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 5
    ws2.column_dimensions["D"].width = 22
    ws2.column_dimensions["E"].width = 22
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 6
    ws2.column_dimensions["H"].width = 5
    ws2.column_dimensions["I"].width = 4
    ws2.column_dimensions["J"].width = 30
    ws2.column_dimensions["K"].width = 40
    ws2.auto_filter.ref = f"A1:K{row-1}"
    ws2.freeze_panes = "A2"

    # ━━━━━━━━━━━━ 시트 3: 코드값 ━━━━━━━━━━━━
    ws3 = wb.create_sheet("코드값")
    ws3.sheet_properties.tabColor = "BF8F00"

    headers3 = ["테이블명", "컬럼명", "코드값", "건수", "비율(%)"]
    for ci, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=ci, value=h)
    style_header(ws3, 1, len(headers3))

    # 코드성 컬럼 판별: 상위 20개 값의 커버리지가 전체의 50% 이상이면 코드 컬럼
    # (날짜/ID 등 고유값이 많은 컬럼 제외)
    NON_CODE_PATTERNS = {"GDAY", "gday", "gamedate", "BIRTH", "INDATE",
                         "GMKEY", "gmkey", "gameID", "GAMEID", "G_ID",
                         "PCODE", "P_ID", "PlayerID", "SERNO", "SEQ_NO",
                         "SeqNO", "REG_DT", "RECORD_DT", "InsertedTime",
                         "INPUTTIME", "LiveText", "LIVETEXT"}

    row = 2
    for domain, tname, tdata in ordered_tables:
        total_rows = tdata["row_count"] or 1
        for col in tdata["columns"]:
            dv = col.get("distinct_values")
            if not dv or len(dv) == 0:
                continue
            cname = col["name"]
            # 명확한 비코드 컬럼 제외
            if cname in NON_CODE_PATTERNS:
                continue
            # 상위 값 커버리지 체크: 상위 20개가 전체의 50% 미만이면 코드 아님
            top_sum = sum(item.get("count", 0) for item in dv)
            if top_sum < total_rows * 0.5:
                continue
            for item in dv:
                val = item.get("value", "")
                cnt = item.get("count", 0)
                pct = round(cnt / total_rows * 100, 1) if total_rows > 0 else 0
                vals = [tname, cname, str(val), cnt, pct]
                for ci, v in enumerate(vals, 1):
                    center = ci in [4, 5]
                    cell = style_cell(ws3, row, ci, center=center)
                    cell.value = v
                row += 1

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 10
    ws3.column_dimensions["E"].width = 8
    ws3.auto_filter.ref = f"A1:E{row-1}"
    ws3.freeze_panes = "A2"

    # ━━━━━━━━━━━━ 시트 4: 문서 정보 ━━━━━━━━━━━━
    ws4 = wb.create_sheet("문서 정보")
    ws4.sheet_properties.tabColor = "7030A0"
    info = [
        ("문서명", "KBO 데이터 표준 정의서 - 데이터 사전"),
        ("RFP ID", "DAR-001"),
        ("생성일", TODAY),
        ("출처", "MSSQL 백업 DB 스키마 직접 추출 (column-metadata.json)"),
        ("테이블 수", f"{len(ordered_tables)}종"),
        ("총 컬럼 수", f"{sum(t['column_count'] for _,_,t in ordered_tables)}개"),
        ("시트 수", "9개 (목차/컬럼사전/코드값/문서정보/표준요약/ID체계/코드사전/용어사전/거버넌스요약)"),
        ("비고", "표준명(안)은 Phase 3 확정 전 초안. 최종 확정 시 갱신 예정."),
    ]
    for ri, (k, v) in enumerate(info, 1):
        ck = ws4.cell(row=ri, column=1, value=k)
        ck.font = Font(name="맑은 고딕", bold=True, size=10)
        cv = ws4.cell(row=ri, column=2, value=v)
        cv.font = Font(name="맑은 고딕", size=10)
    ws4.column_dimensions["A"].width = 15
    ws4.column_dimensions["B"].width = 60

    # ━━━━━━━━━━━━ 시트 5: 표준 요약 ━━━━━━━━━━━━
    ws5 = wb.create_sheet("표준 요약")
    ws5.sheet_properties.tabColor = "00B050"

    STD_SUMMARY = [
        # (구분, 계층, 규칙, 예시)
        ("4계층 명명", "DB 테이블", "UPPER_SNAKE_CASE", "GAME_INFO"),
        ("4계층 명명", "DB 컬럼", "lower_snake_case + 접미사", "game_id, avg_rt"),
        ("4계층 명명", "API URL", "kebab-case", "/game-info/{gameId}"),
        ("4계층 명명", "API 필드", "camelCase", "gameId, avgRt"),
        ("4계층 명명", "Kafka 토픽", "dot.separated.lower", "kbo.game.play"),
        ("4계층 명명", "Kafka 필드", "camelCase", "gameId"),
        ("4계층 명명", "WebSocket 경로", "path/lower", "/ws/game/{gameId}"),
        ("4계층 명명", "WebSocket 필드", "camelCase", "gameId"),
        ("도메인 타입", "_id (식별자)", "char/int (ID별 상이)", "game_id char(13), player_id int"),
        ("도메인 타입", "_nm (이름)", "nvarchar(100/500)", "player_nm, stadium_nm"),
        ("도메인 타입", "_cd (코드)", "varchar(10) / char(2)", "how_cd, team_cd"),
        ("도메인 타입", "_sc (상태코드)", "varchar(10) / char(2)", "top_bottom_sc, wls_sc"),
        ("도메인 타입", "_cn (수량)", "int NOT NULL DEFAULT 0", "hit_cn, hr_cn"),
        ("도메인 타입", "_rt (비율)", "decimal(8,5)", "avg_rt, era_rt"),
        ("도메인 타입", "_dt (날짜)", "char(8) / datetime2", "game_dt, reg_dt"),
        ("도메인 타입", "_tm (시각)", "char(4) / int", "start_tm, game_duration_tm"),
        ("도메인 타입", "_if (불리언)", "bit NOT NULL DEFAULT 0", "first_if, cancel_if"),
        ("도메인 타입", "_va (측정값)", "int / decimal", "temperature_va, height_va"),
        ("도메인 타입", "_no (순번)", "int", "inning_no, seq_no"),
        ("접미사→MSSQL", "_id → char/int", "식별자별 타입 상이", "game_id=char(13), player_id=int"),
        ("접미사→MSSQL", "_nm → nvarchar", "한글=nvarchar 필수", "varchar 사용 금지"),
        ("접미사→MSSQL", "_cn → int", "NOT NULL DEFAULT 0", "NULL 수량=집계 오류"),
        ("접미사→MSSQL", "_rt → decimal(8,5)", "float 금지", "부동소수점 노이즈 방지"),
        ("변환 5원칙", "1. 접미사 부여", "모든 컬럼에 타입 접미사", "PCODE → player_id"),
        ("변환 5원칙", "2. snake_case", "camelCase/PascalCase → snake", "BatTotal → bat_total"),
        ("변환 5원칙", "3. 한글→nvarchar", "varchar 한글 컬럼 전환", "NAME varchar→player_nm nvarchar"),
        ("변환 5원칙", "4. float→decimal", "비율 컬럼 정밀도 확보", "HRA float→avg_rt decimal(8,5)"),
        ("변환 5원칙", "5. ID 통일", "구세대→신세대 명칭 통일", "GMKEY/G_ID→game_id"),
    ]
    headers5 = ["구분", "계층/항목", "규칙", "예시"]
    for ci, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=ci, value=h)
    style_header(ws5, 1, len(headers5))

    for ri, (cat, layer, rule, ex) in enumerate(STD_SUMMARY, 2):
        for ci, v in enumerate([cat, layer, rule, ex], 1):
            cell = style_cell(ws5, ri, ci, center=(ci == 1))
            cell.value = v

    ws5.column_dimensions["A"].width = 16
    ws5.column_dimensions["B"].width = 24
    ws5.column_dimensions["C"].width = 36
    ws5.column_dimensions["D"].width = 44
    ws5.auto_filter.ref = f"A1:D{len(STD_SUMMARY)+1}"
    ws5.freeze_panes = "A2"

    # ━━━━━━━━━━━━ 시트 6: ID 체계 ━━━━━━━━━━━━
    ws6 = wb.create_sheet("ID 체계")
    ws6.sheet_properties.tabColor = "FF6600"

    ID_SYSTEM = [
        # (ID명, 레거시명, 타입, 형식, 예시, 비고)
        ("game_id", "GMKEY, G_ID, gmkey, gameID", "char(13)", "YYYYMMDDVVHH#", "20250322HHKT0", "경기 고유키. 팀코드 역사적 고정"),
        ("player_id", "PCODE, P_ID, PlayerID", "int", "5자리 숫자", "75847", "레거시 varchar(10)→int 전환"),
        ("team_id", "T_ID, TEAM, HTEAM, VTEAM", "char(2)", "영문 대문자 2자", "HH (키움)", "구단명 변경에도 코드 불변 (OB=두산)"),
        ("stadium_id", "stadium_key, STAD", "char(2)", "영문 대문자 2자", "JS (잠실)", "구장명 변경에도 코드 불변"),
        ("season_id", "GYEAR, SEASON_ID, gyear", "smallint", "4자리 연도", "2025, 9999=통산", "9999는 Career Total 예약값"),
        ("league_id", "LE_ID", "smallint", "정수", "1=정규, 2=가을리그", ""),
        ("series_id", "SR_ID", "smallint", "정수", "0=정규시즌, 1=포스트", "S2i 확인 필요 (1~9)"),
        ("", "", "", "", "", ""),
        ("── 복합 PK ──", "", "", "", "", ""),
        ("GAME_INFO", "", "", "game_id", "", "경기 1건=1행"),
        ("HITTER", "", "", "game_id + season_yr + player_id", "", "경기별 타자 기록"),
        ("PITCHER", "", "", "game_id + season_yr + player_id", "", "경기별 투수 기록"),
        ("GAME_CONT_APP", "", "", "game_id + season_yr + seq_no", "", "경기 이벤트 상세"),
        ("ENTRY", "", "", "game_id + season_yr + turn_no + player_id + position_cd", "", "출전 엔트리"),
        ("BAT_TOTAL", "", "", "season_yr + player_id + series_cd", "", "타격 시즌/통산"),
        ("PIT_TOTAL", "", "", "season_yr + player_id + series_cd", "", "투구 시즌/통산"),
        ("SEASON_PLAYER_HITTER", "", "", "season_id + player_id + section_cd + group_if", "", "시즌 타자 통계"),
        ("SCORE", "", "", "game_id + season_yr + top_bottom_cd", "", "팀별 경기 스코어"),
        ("GAME_HR", "", "", "league_id + series_id + game_id + seq_no", "", "홈런 기록"),
        ("PERSON", "", "", "season_yr + player_id", "", "시즌별 선수 정보"),
    ]
    headers6 = ["ID명/테이블", "레거시명", "타입", "형식/PK 구성", "예시", "비고"]
    for ci, h in enumerate(headers6, 1):
        ws6.cell(row=1, column=ci, value=h)
    style_header(ws6, 1, len(headers6))

    for ri, row_data in enumerate(ID_SYSTEM, 2):
        for ci, v in enumerate(row_data, 1):
            cell = style_cell(ws6, ri, ci)
            cell.value = v
            if row_data[0].startswith("──"):
                cell.font = Font(name="맑은 고딕", bold=True, size=9)

    ws6.column_dimensions["A"].width = 28
    ws6.column_dimensions["B"].width = 32
    ws6.column_dimensions["C"].width = 12
    ws6.column_dimensions["D"].width = 44
    ws6.column_dimensions["E"].width = 28
    ws6.column_dimensions["F"].width = 36
    ws6.freeze_panes = "A2"

    # ━━━━━━━━━━━━ 시트 7: 코드 사전 ━━━━━━━━━━━━
    ws7 = wb.create_sheet("코드 사전")
    ws7.sheet_properties.tabColor = "C00000"

    CODE_DICT = [
        # (분류, 코드명, 코드값, 한글명, 영문명, 비고)
        # TB
        ("경기 이벤트", "top_bottom_cd", "T", "원정 (초공)", "Top", "이닝 앞부분 공격"),
        ("경기 이벤트", "top_bottom_cd", "B", "홈 (말공)", "Bottom", "이닝 뒷부분 공격"),
        # how_cd 타격
        ("타격 결과", "how_cd", "H1", "1루타", "Single", ""),
        ("타격 결과", "how_cd", "H2", "2루타", "Double", ""),
        ("타격 결과", "how_cd", "H3", "3루타", "Triple", ""),
        ("타격 결과", "how_cd", "HR", "홈런", "Home Run", ""),
        ("타격 결과", "how_cd", "HI", "내야안타", "Infield Hit", ""),
        ("타격 결과", "how_cd", "HB", "번트안타", "Bunt Hit", ""),
        ("타격 결과", "how_cd", "BH", "기타안타", "Bunt Hit (alt?)", "S2i 확인 필요 (14,749건)"),
        # how_cd 아웃
        ("아웃 결과", "how_cd", "GR", "땅볼아웃", "Ground Out", ""),
        ("아웃 결과", "how_cd", "FL", "플라이아웃", "Fly Out", ""),
        ("아웃 결과", "how_cd", "KK", "삼진 (헛스윙)", "Strikeout Swinging", ""),
        ("아웃 결과", "how_cd", "KN", "삼진 (루킹)", "Strikeout Called", ""),
        ("아웃 결과", "how_cd", "FO", "파울플라이아웃", "Foul Out", ""),
        ("아웃 결과", "how_cd", "FF", "파울플라이", "Foul Fly", ""),
        ("아웃 결과", "how_cd", "GD", "병살타", "Grounded into DP", ""),
        ("아웃 결과", "how_cd", "IF", "인필드플라이", "Infield Fly", ""),
        ("아웃 결과", "how_cd", "FD", "파울 병살", "Foul DP", ""),
        ("아웃 결과", "how_cd", "FC", "야수선택", "Fielder's Choice", ""),
        ("아웃 결과", "how_cd", "FE", "실책 출루", "Error (reached)", ""),
        ("아웃 결과", "how_cd", "ER", "실책", "Error", ""),
        ("아웃 결과", "how_cd", "LN", "라인드라이브아웃", "Line Drive Out", ""),
        ("아웃 결과", "how_cd", "SF", "희생플라이", "Sacrifice Fly", ""),
        ("아웃 결과", "how_cd", "SH", "희생번트", "Sacrifice Hit", ""),
        ("아웃 결과", "how_cd", "B2", "번트 병살", "Bunt DP", ""),
        # how_cd 볼넷/사구
        ("볼넷/사구", "how_cd", "BB", "볼넷", "Base on Balls", ""),
        ("볼넷/사구", "how_cd", "IB", "고의사구", "Intentional Walk", ""),
        ("볼넷/사구", "how_cd", "HP", "사구", "Hit by Pitch", ""),
        # how_cd 주루
        ("주루", "how_cd", "SB", "도루", "Stolen Base", ""),
        ("주루", "how_cd", "CS", "도루실패", "Caught Stealing", ""),
        ("주루", "how_cd", "WP", "폭투", "Wild Pitch", ""),
        ("주루", "how_cd", "PB", "포일", "Passed Ball", ""),
        ("주루", "how_cd", "BK", "보크", "Balk", ""),
        ("주루", "how_cd", "BN", "보크 진루", "Balk (advance)", ""),
        ("주루", "how_cd", "OB", "주루방해", "Obstruction", ""),
        # how_cd 기타
        ("기타 이벤트", "how_cd", "KB", "삼진+폭투/포일 출루", "K+WP/PB", ""),
        ("기타 이벤트", "how_cd", "KP", "삼진+폭투", "K+WP", ""),
        ("기타 이벤트", "how_cd", "KW", "삼진+폭투/야선", "K+WP/FC", ""),
        ("기타 이벤트", "how_cd", "LF", "라인드라이브 파울", "Line Drive Foul", ""),
        ("기타 이벤트", "how_cd", "TH", "송구", "Throw", ""),
        ("기타 이벤트", "how_cd", "PR", "대주자", "Pinch Runner", ""),
        # place_cd
        ("타구 방향", "place_cd", "0", "해당없음", "N/A", "볼넷/삼진 등"),
        ("타구 방향", "place_cd", "1", "투수 방면", "Pitcher", ""),
        ("타구 방향", "place_cd", "2", "포수 방면", "Catcher", ""),
        ("타구 방향", "place_cd", "3", "1루수 방면", "First Base", ""),
        ("타구 방향", "place_cd", "4", "2루수 방면", "Second Base", ""),
        ("타구 방향", "place_cd", "5", "3루수 방면", "Third Base", ""),
        ("타구 방향", "place_cd", "6", "유격수 방면", "Shortstop", ""),
        ("타구 방향", "place_cd", "7", "좌익수 방면", "Left Field", ""),
        ("타구 방향", "place_cd", "8", "중견수 방면", "Center Field", ""),
        ("타구 방향", "place_cd", "9", "우익수 방면", "Right Field", ""),
        ("타구 방향", "place_cd", "78", "좌중간", "Left-Center", "GAME_HR"),
        ("타구 방향", "place_cd", "89", "우중간", "Right-Center", "GAME_HR"),
        # wls_cd
        ("승패세", "wls_cd", "W", "승리", "Win", ""),
        ("승패세", "wls_cd", "L", "패배", "Loss", ""),
        ("승패세", "wls_cd", "S", "세이브", "Save", ""),
        # 팀 코드
        ("팀 (현행)", "team_id", "HH", "키움 히어로즈", "Kiwoom Heroes", "서울(고척)"),
        ("팀 (현행)", "team_id", "HT", "KIA 타이거즈", "KIA Tigers", "광주"),
        ("팀 (현행)", "team_id", "KT", "KT 위즈", "KT Wiz", "수원"),
        ("팀 (현행)", "team_id", "LG", "LG 트윈스", "LG Twins", "서울(잠실)"),
        ("팀 (현행)", "team_id", "LT", "롯데 자이언츠", "Lotte Giants", "부산"),
        ("팀 (현행)", "team_id", "NC", "NC 다이노스", "NC Dinos", "창원"),
        ("팀 (현행)", "team_id", "OB", "두산 베어스", "Doosan Bears", "서울(잠실)"),
        ("팀 (현행)", "team_id", "SK", "SSG 랜더스", "SSG Landers", "인천"),
        ("팀 (현행)", "team_id", "SS", "삼성 라이온즈", "Samsung Lions", "대구"),
        ("팀 (현행)", "team_id", "WO", "한화 이글스", "Hanwha Eagles", "대전"),
        ("팀 (역사)", "team_id", "HD", "현대(삼미→청보→태평양→현대)", "Hyundai", "1982~2007 해체"),
        ("팀 (역사)", "team_id", "SB", "쌍방울", "Ssangbangwool", "1991~1999 해체"),
        # 구장
        ("구장", "stadium_id", "JS", "잠실야구장", "Jamsil", "LG, OB"),
        ("구장", "stadium_id", "KC", "고척스카이돔", "Gocheok", "HH"),
        ("구장", "stadium_id", "GJ", "광주-KIA챔피언스필드", "Gwangju", "HT"),
        ("구장", "stadium_id", "SW", "수원KT위즈파크", "Suwon", "KT"),
        ("구장", "stadium_id", "DG", "대구삼성라이온즈파크", "Daegu", "SS"),
        ("구장", "stadium_id", "DJ", "대전한화생명이글스파크", "Daejeon", "WO"),
        ("구장", "stadium_id", "IC", "인천SSG랜더스필드", "Incheon", "SK"),
        ("구장", "stadium_id", "SJ", "사직야구장", "Sajik", "LT"),
        ("구장", "stadium_id", "CW", "창원NC파크", "Changwon", "NC"),
        ("구장", "stadium_id", "UL", "울산문수야구장", "Ulsan", ""),
        ("구장", "stadium_id", "PH", "포항야구장", "Pohang", ""),
        ("구장", "stadium_id", "CJ", "청주야구장", "Cheongju", ""),
        ("구장", "stadium_id", "CH", "춘천야구장", "Chuncheon", ""),
        ("구장", "stadium_id", "JJ", "전주야구장", "Jeonju", ""),
        ("구장", "stadium_id", "JE", "제주야구장", "Jeju", ""),
        ("구장", "stadium_id", "MS", "마산야구장", "Masan", ""),
        ("구장", "stadium_id", "DM", "동대문야구장", "Dongdaemun", "폐장"),
        ("구장", "stadium_id", "KD", "구덕야구장", "Gudeok", "폐장"),
        # 날씨
        ("날씨", "weather_cd", "F", "맑음", "Fine", ""),
        ("날씨", "weather_cd", "C", "흐림", "Cloudy", ""),
        ("날씨", "weather_cd", "R", "비", "Rain", ""),
        ("날씨", "weather_cd", "CR", "흐리고 비", "Cloudy then Rain", ""),
        ("날씨", "weather_cd", "FC", "맑다가 흐림", "Fine then Cloudy", ""),
        ("날씨", "weather_cd", "CF", "흐리다가 맑음", "Cloudy then Fine", ""),
        # 방송
        ("방송사", "broadcasting_cd", "0", "비해당", "N/A", ""),
        ("방송사", "broadcasting_cd", "1", "SBS", "SBS", ""),
        ("방송사", "broadcasting_cd", "2", "KBS2", "KBS2", ""),
        ("방송사", "broadcasting_cd", "3", "KBS1", "KBS1", ""),
        ("방송사", "broadcasting_cd", "4", "MBC", "MBC", ""),
        ("방송사", "broadcasting_cd", "5", "SBS SPORTS", "SBS SPORTS", ""),
        ("방송사", "broadcasting_cd", "6", "KBSN스포츠", "KBSN SPORTS", ""),
        ("방송사", "broadcasting_cd", "7", "MBC SPORTS+", "MBC SPORTS+", ""),
        ("방송사", "broadcasting_cd", "8", "SPOTV", "SPOTV", ""),
        ("방송사", "broadcasting_cd", "9", "SPOTV2", "SPOTV2", ""),
        ("방송사", "broadcasting_cd", "10", "TVING", "TVING", ""),
        ("방송사", "broadcasting_cd", "11", "tvN SPORTS", "tvN SPORTS", ""),
        # 경기 상태
        ("경기 상태", "game_sc_id", "70", "종료", "Final", "확인됨"),
        ("경기 상태", "game_sc_id", "10", "예정", "Scheduled", "추정"),
        ("경기 상태", "game_sc_id", "20", "진행중", "In Progress", "추정"),
        ("경기 상태", "game_sc_id", "50", "취소", "Cancelled", "추정"),
        # 경기 유형
        ("경기 유형", "game_type_cd", "0", "정규시즌", "Regular Season", ""),
        ("경기 유형", "game_type_cd", "1", "포스트시즌", "Postseason", ""),
        ("경기 유형", "game_type_cd", "3", "올스타전", "All-Star Game", "추정"),
        ("경기 유형", "game_type_cd", "5", "시범경기", "Exhibition", "추정"),
        # 기록 상태 (신규)
        ("기록 상태", "record_status_cd", "DRAFT", "입력", "Draft", "기록원 실시간 입력"),
        ("기록 상태", "record_status_cd", "REVIEW", "검증", "Under Review", "기록위원 검증 중"),
        ("기록 상태", "record_status_cd", "CONFIRMED", "확정", "Confirmed", "공식 기록 확정"),
        ("기록 상태", "record_status_cd", "REVISED", "정정", "Revised", "확정 후 오류 정정"),
        # 포지션
        ("포지션", "position Y", "1", "투수", "Pitcher (P)", ""),
        ("포지션", "position Y", "2", "포수", "Catcher (C)", ""),
        ("포지션", "position Y", "3", "1루수", "First Base (1B)", ""),
        ("포지션", "position Y", "4", "2루수", "Second Base (2B)", ""),
        ("포지션", "position Y", "5", "3루수", "Third Base (3B)", ""),
        ("포지션", "position Y", "6", "유격수", "Shortstop (SS)", ""),
        ("포지션", "position Y", "7", "좌익수", "Left Field (LF)", ""),
        ("포지션", "position Y", "8", "중견수", "Center Field (CF)", ""),
        ("포지션", "position Y", "9", "우익수", "Right Field (RF)", ""),
        ("포지션", "position Y", "D", "지명타자", "Designated Hitter", ""),
        ("포지션", "position Y", "H", "대타", "Pinch Hitter", ""),
        ("포지션", "position Y", "R", "대주자", "Pinch Runner", ""),
    ]
    headers7 = ["분류", "코드명", "코드값", "한글명", "영문명", "비고"]
    for ci, h in enumerate(headers7, 1):
        ws7.cell(row=1, column=ci, value=h)
    style_header(ws7, 1, len(headers7))

    for ri, row_data in enumerate(CODE_DICT, 2):
        for ci, v in enumerate(row_data, 1):
            cell = style_cell(ws7, ri, ci, center=(ci == 3))
            cell.value = v

    ws7.column_dimensions["A"].width = 14
    ws7.column_dimensions["B"].width = 20
    ws7.column_dimensions["C"].width = 12
    ws7.column_dimensions["D"].width = 24
    ws7.column_dimensions["E"].width = 24
    ws7.column_dimensions["F"].width = 30
    ws7.auto_filter.ref = f"A1:F{len(CODE_DICT)+1}"
    ws7.freeze_panes = "A2"

    # ━━━━━━━━━━━━ 시트 8: 용어 사전 ━━━━━━━━━━━━
    ws8 = wb.create_sheet("용어 사전")
    ws8.sheet_properties.tabColor = "0070C0"

    GLOSSARY = [
        # (도메인, 용어, 영문, 약어, 정의, DB컬럼, 관련 테이블)
        # 타격
        ("야구 기록-타격", "타석", "Plate Appearance", "PA", "타자가 타석에 들어선 횟수", "pa", "Hitter, BatTotal"),
        ("야구 기록-타격", "타수", "At Bat", "AB", "볼넷·사구·희생 제외 횟수", "ab", "Hitter, BatTotal"),
        ("야구 기록-타격", "안타", "Hit", "HIT", "안전하게 출루한 타격", "hit", "Hitter, BatTotal"),
        ("야구 기록-타격", "1루타", "Single", "H1", "1루까지 진루한 안타", "h1b", "Hitter"),
        ("야구 기록-타격", "2루타", "Double", "H2", "2루까지 진루한 안타", "h2b", "Hitter, BatTotal"),
        ("야구 기록-타격", "3루타", "Triple", "H3", "3루까지 진루한 안타", "h3b", "Hitter, BatTotal"),
        ("야구 기록-타격", "홈런", "Home Run", "HR", "본루까지 진루한 안타", "hr", "Hitter, BatTotal, GAME_HR"),
        ("야구 기록-타격", "타점", "Runs Batted In", "RBI", "타자로 인한 득점 수", "rbi", "Hitter, BatTotal"),
        ("야구 기록-타격", "득점", "Run Scored", "RUN", "본루를 밟아 득점한 횟수", "run", "Hitter, BatTotal"),
        ("야구 기록-타격", "볼넷", "Base on Balls", "BB", "4볼로 1루 출루", "bb", "Hitter, Pitcher"),
        ("야구 기록-타격", "사구", "Hit by Pitch", "HBP", "몸에 맞아 1루 출루", "hbp", "Hitter, BatTotal"),
        ("야구 기록-타격", "고의사구", "Intentional Walk", "IBB", "투수가 의도적으로 볼넷 허용", "ibb", "Hitter, BatTotal"),
        ("야구 기록-타격", "삼진", "Strikeout", "SO", "3스트라이크로 아웃", "so", "Hitter, Pitcher"),
        ("야구 기록-타격", "도루", "Stolen Base", "SB", "투구 중 다음 루로 진루", "sb", "Hitter, BatTotal"),
        ("야구 기록-타격", "도루실패", "Caught Stealing", "CS", "도루 시도 중 태그아웃", "cs", "Hitter, BatTotal"),
        ("야구 기록-타격", "희생번트", "Sacrifice Hit", "SH", "번트로 주자 진루", "sh", "Hitter, BatTotal"),
        ("야구 기록-타격", "희생플라이", "Sacrifice Fly", "SF", "플라이로 3루 주자 득점", "sf", "Hitter, BatTotal"),
        ("야구 기록-타격", "병살타", "Grounded into DP", "GIDP", "타격으로 2명 동시 아웃", "gidp", "Hitter, BatTotal"),
        ("야구 기록-타격", "실책", "Error", "ERR", "수비 실수로 출루/진루", "err", "Hitter, DEFEN"),
        ("야구 기록-타격", "잔루", "Left on Base", "LOB", "이닝 종료 시 잔류 주자 수", "lob", "Hitter"),
        # 투수
        ("야구 기록-투수", "이닝", "Innings Pitched", "IP", "투수가 던진 이닝 수", "ip", "Pitcher, PitTotal"),
        ("야구 기록-투수", "상대타자수", "Batters Faced", "BF", "상대한 총 타자 수", "bf", "Pitcher, PitTotal"),
        ("야구 기록-투수", "승리", "Win", "W", "승리 투수 기록", "win", "Pitcher, PitTotal"),
        ("야구 기록-투수", "패배", "Loss", "L", "패전 투수 기록", "loss", "Pitcher, PitTotal"),
        ("야구 기록-투수", "세이브", "Save", "SV", "리드 지키며 마무리", "sv", "Pitcher, PitTotal"),
        ("야구 기록-투수", "홀드", "Hold", "HLD", "리드 유지 구원 투수", "hld", "Pitcher, PitTotal"),
        ("야구 기록-투수", "블론세이브", "Blown Save", "BS", "세이브 상황 리드 반납", "bs", "Pitcher, PitTotal"),
        ("야구 기록-투수", "완투", "Complete Game", "CG", "경기 전체를 혼자 투구", "cg", "Pitcher, PitTotal"),
        ("야구 기록-투수", "완봉", "Shutout", "SHO", "완투하며 무실점", "sho", "Pitcher, PitTotal"),
        ("야구 기록-투수", "자책점", "Earned Run", "ER", "투수 책임 실점", "er", "Pitcher, PitTotal"),
        ("야구 기록-투수", "실점", "Runs Allowed", "R", "등판 중 상대 총 득점", "runs_cn", "Pitcher, PitTotal"),
        ("야구 기록-투수", "투구수", "Number of Pitches", "NP", "던진 총 투구 수", "pitch_cn", "Pitcher"),
        ("야구 기록-투수", "폭투", "Wild Pitch", "WP", "포수 못 잡는 투구", "wp", "Pitcher, PitTotal"),
        ("야구 기록-투수", "보크", "Balk", "BK", "규칙 위반 투구 동작", "bk", "Pitcher, PitTotal"),
        # 수비
        ("야구 기록-수비", "자살", "Put Out", "PO", "직접 아웃 기록", "po", "DEFEN"),
        ("야구 기록-수비", "보살", "Assist", "AST", "아웃 도움", "ast", "DEFEN"),
        ("야구 기록-수비", "병살", "Double Play", "DP", "1플레이 2아웃", "dp", "DEFEN"),
        # 비율
        ("세이버메트릭스", "타율", "Batting Average", "AVG", "안타/타수", "avg_rt", "BatTotal"),
        ("세이버메트릭스", "출루율", "On-Base Percentage", "OBP", "(안타+볼넷+사구)/(타수+볼넷+사구+희비)", "obp_rt", "SEASON_PLAYER_HITTER"),
        ("세이버메트릭스", "장타율", "Slugging Percentage", "SLG", "루타/타수", "slg_rt", "SEASON_PLAYER_HITTER"),
        ("세이버메트릭스", "OPS", "OBP + SLG", "OPS", "출루율+장타율", "ops_rt", "SEASON_PLAYER_HITTER"),
        ("세이버메트릭스", "순장타율", "Isolated Power", "ISO", "장타율-타율", "iso_rt", "SEASON_PLAYER_HITTER"),
        ("세이버메트릭스", "BABIP", "Batting Avg on BIP", "BABIP", "인플레이 타구 안타 비율", "babip_rt", "SEASON_PLAYER_HITTER"),
        ("세이버메트릭스", "평균자책점", "Earned Run Average", "ERA", "9이닝당 자책점", "era_rt", "PitTotal"),
        ("세이버메트릭스", "WHIP", "Walks+Hits per IP", "WHIP", "이닝당 볼넷+안타", "whip_rt", "SEASON_PLAYER_PITCHER"),
        ("세이버메트릭스", "피안타율", "Opponent AVG", "OAVG", "상대 타자 타율", "opp_avg_rt", "SEASON_PLAYER_PITCHER"),
        # 경기 운영
        ("경기 운영", "경기키", "Game Key", "-", "13자리 경기 고유 식별 코드", "game_id", "전체"),
        ("경기 운영", "경기일자", "Game Date", "-", "경기 날짜 (YYYYMMDD)", "game_dt", "GAMEINFO"),
        ("경기 운영", "시즌 연도", "Season Year", "-", "4자리 연도. 9999=통산", "season_yr", "전체"),
        ("경기 운영", "더블헤더", "Doubleheader", "DH", "같은 날 같은 팀 2경기", "doubleheader_no", "GAMEINFO"),
        ("경기 운영", "관중수", "Attendance", "-", "입장 관중 수", "crowd_cn", "GAMEINFO"),
        ("경기 운영", "초/말", "Top/Bottom", "TB", "이닝 전반(T)/후반(B)", "top_bottom_cd", "Hitter, Pitcher"),
        # 선수/팀
        ("선수/팀", "선수 코드", "Player Code", "-", "선수 고유 식별 코드 (5자리)", "player_id", "전체"),
        ("선수/팀", "선수명", "Player Name", "-", "한글 이름", "player_nm", "person"),
        ("선수/팀", "팀 코드", "Team Code", "-", "팀 식별 2자리 코드", "team_id", "전체"),
        ("선수/팀", "등번호", "Back Number", "-", "유니폼 번호", "back_no", "person"),
        ("선수/팀", "포지션", "Position", "POS", "수비 위치", "position_cd", "person, ENTRY"),
        # 데이터 품질
        ("데이터 품질", "EUC-KR 인코딩", "-", "-", "레거시 varchar 한글 인코딩. nvarchar 환경에서 깨짐", "-", "-"),
        ("데이터 품질", "-1 센티널", "-", "-", "Score 이닝에서 -1=미진행 이닝", "-", "Score"),
        ("데이터 품질", "9999 예약값", "-", "-", "GYEAR=9999는 통산 기록", "-", "BatTotal, PitTotal"),
        ("데이터 품질", "T/B 합계행", "-", "-", "PCODE='T'/'B'는 팀 합계행", "-", "Hitter, Pitcher"),
    ]
    headers8 = ["도메인", "용어", "영문", "약어", "정의", "DB 컬럼", "관련 테이블"]
    for ci, h in enumerate(headers8, 1):
        ws8.cell(row=1, column=ci, value=h)
    style_header(ws8, 1, len(headers8))

    for ri, row_data in enumerate(GLOSSARY, 2):
        for ci, v in enumerate(row_data, 1):
            cell = style_cell(ws8, ri, ci)
            cell.value = v

    ws8.column_dimensions["A"].width = 18
    ws8.column_dimensions["B"].width = 14
    ws8.column_dimensions["C"].width = 24
    ws8.column_dimensions["D"].width = 8
    ws8.column_dimensions["E"].width = 38
    ws8.column_dimensions["F"].width = 16
    ws8.column_dimensions["G"].width = 30
    ws8.auto_filter.ref = f"A1:G{len(GLOSSARY)+1}"
    ws8.freeze_panes = "A2"

    # ━━━━━━━━━━━━ 시트 9: 거버넌스 요약 ━━━━━━━━━━━━
    ws9 = wb.create_sheet("거버넌스 요약")
    ws9.sheet_properties.tabColor = "404040"

    GOV_SUMMARY = [
        # (구분, 항목, 내용1, 내용2, 내용3)
        ("역할 정의", "R-01 S2i 데이터 담당", "Sports2i (외주)", "DB2 데이터 전송, 스키마 변경 통보", "DB2: W"),
        ("역할 정의", "R-02 기록원 (Scorer)", "KBO 기록위원회", "경기 중 실시간 입력 (DRAFT)", "DB2: R"),
        ("역할 정의", "R-03 기록위원회", "KBO 기록위원회", "기록 검증(REVIEW)/확정(CONFIRMED)", "DB1: R"),
        ("역할 정의", "R-04 DBA", "KBO IT팀", "DB 스키마 관리, 백업, 접근 권한", "전체: RWA"),
        ("역할 정의", "R-05 거버넌스 위원회", "KBO IT+기록위+운영", "표준 문서 개정, 코드 확정, 최종 승인", "문서: E/A"),
        ("역할 정의", "R-06 데이터팀", "KBO IT팀", "KBO 자체 테이블, 세이버, 품질 모니터링", "KBO DB: RW"),
        ("", "", "", "", ""),
        ("오너십 매트릭스", "경기 기록 (game/) 12종", "생성: R-01 (S2i)", "품질: R-03 (기록위)", "스키마: R-05"),
        ("오너십 매트릭스", "통계 (stats/) 10종", "생성: R-01 (S2i)", "품질: R-03 (기록위)", "스키마: R-05"),
        ("오너십 매트릭스", "실시간 (realtime/) 9종", "생성: R-01 (S2i)", "품질: R-01 (S2i)", "스키마: R-05"),
        ("오너십 매트릭스", "마스터 (master/) 8종", "생성: R-01/R-04", "품질: R-04 (DBA)", "스키마: R-05"),
        ("오너십 매트릭스", "KBO 자체 (신규)", "생성: R-06 (데이터팀)", "품질: R-06 (데이터팀)", "스키마: R-05"),
        ("", "", "", "", ""),
        ("품질 KPI", "1. 완결성 (Completeness)", "1-(NULL PK행/전체행)", "목표: 100%", "일간, R-04"),
        ("품질 KPI", "2. 오류율 (Error Rate)", "CRITICAL 위반/전체행", "목표: <0.01%", "일간, R-04"),
        ("품질 KPI", "3. 인코딩 정상률", "nvarchar 정상행/한글행", "목표: >99.9%", "주간, R-06"),
        ("품질 KPI", "4. FK 불일치율", "FK실패행/자식테이블행", "목표: <0.1%", "주간, R-06"),
        ("품질 KPI", "5. 미확정 경기 수", "record_status!=CONFIRMED", "목표: 0건(전일이전)", "일간, R-03"),
        ("", "", "", "", ""),
        ("변경 유형", "C1: 코드값 추가", "영향: 낮음", "승인: R-05", "3영업일"),
        ("변경 유형", "C2: 코드값 의미 변경", "영향: 중간", "승인: R-05", "5영업일"),
        ("변경 유형", "C3: 미확인 코드 확정", "영향: 중간", "승인: R-05", "5영업일"),
        ("변경 유형", "S1: 컬럼 추가", "영향: 중간", "승인: R-05", "5영업일"),
        ("변경 유형", "S2: 컬럼 삭제 (비활성화)", "영향: 높음", "승인: R-05", "10영업일"),
        ("변경 유형", "S3: 테이블 신규 생성", "영향: 높음", "승인: R-05", "10영업일"),
        ("변경 유형", "S4: S2i 전송 스키마 변경", "영향: 높음", "승인: R-05+R-01", "협의"),
        ("변경 유형", "D1: 표준 문서 오탈자", "영향: 낮음", "승인: R-06 자체", "1영업일"),
        ("변경 유형", "D2: 표준 문서 내용 변경", "영향: 중간", "승인: R-05", "5영업일"),
        ("변경 유형", "D3: 표준 문서 구조 개편", "영향: 높음", "승인: R-05", "10영업일"),
    ]
    headers9 = ["구분", "항목", "내용 1", "내용 2", "내용 3"]
    for ci, h in enumerate(headers9, 1):
        ws9.cell(row=1, column=ci, value=h)
    style_header(ws9, 1, len(headers9))

    for ri, row_data in enumerate(GOV_SUMMARY, 2):
        for ci, v in enumerate(row_data, 1):
            cell = style_cell(ws9, ri, ci)
            cell.value = v
            if row_data[0] and ci == 1:
                cell.font = Font(name="맑은 고딕", bold=True, size=9)

    ws9.column_dimensions["A"].width = 18
    ws9.column_dimensions["B"].width = 32
    ws9.column_dimensions["C"].width = 28
    ws9.column_dimensions["D"].width = 30
    ws9.column_dimensions["E"].width = 20
    ws9.freeze_panes = "A2"

    # ── 저장 ──
    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"생성 완료: {OUT_PATH}")
    print(f"  - 시트 1 목차: {len(ordered_tables)}개 테이블")
    total_cols = sum(t["column_count"] for _, _, t in ordered_tables)
    print(f"  - 시트 2 컬럼 사전: {total_cols}개 컬럼")
    print(f"  - 시트 3 코드값: {row-2}개 항목")
    print(f"  - 시트 5 표준 요약: {len(STD_SUMMARY)}개 항목")
    print(f"  - 시트 6 ID 체계: {len(ID_SYSTEM)}개 항목")
    print(f"  - 시트 7 코드 사전: {len(CODE_DICT)}개 항목")
    print(f"  - 시트 8 용어 사전: {len(GLOSSARY)}개 항목")
    print(f"  - 시트 9 거버넌스 요약: {len(GOV_SUMMARY)}개 항목")
    print(f"  - 생성일: {TODAY}")


if __name__ == "__main__":
    main()
