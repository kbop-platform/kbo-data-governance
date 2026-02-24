#!/usr/bin/env python3
"""
compare-excel-vs-db.py

Compares the Excel table inventory (S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx)
with the actual backup DB schema JSON files in raw/.

Outputs:
  1. Per-sheet summary of all tables listed in Excel
  2. Per-JSON-file summary of all tables in our backup DB
  3. Cross-comparison: tables in Excel vs. backup DB
"""

import os
import json
import glob
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx")
RAW_DIR = os.path.join(BASE_DIR, "raw")

SHEET_NAMES = ["DB1_", "DB2_", "전송 외"]

# ──────────────────────────────────────────────
# 1. Read Excel
# ──────────────────────────────────────────────

def read_excel_sheet(ws):
    """
    Read a single sheet. Returns list of dicts with:
      db_name, no, table, col_count, row_count,
      s2i_db, s2i_table, s2i_col_count, s2i_row_count,
      col_match, row_match, etc, row_reason
    """
    rows = []
    current_db = None

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        # row is a tuple; columns indexed 0..12
        db_name = row[0]
        no = row[1]
        table = row[2]
        col_count = row[3]
        row_count = row[4]
        s2i_db = row[5]
        s2i_table = row[6]
        s2i_col_count = row[7]
        s2i_row_count = row[8]
        col_match = row[9]
        row_match = row[10]
        etc = row[11]
        row_reason = row[12]

        # Forward-fill DB name from merged cells
        if db_name is not None and str(db_name).strip():
            current_db = str(db_name).strip()

        # Skip rows where TABLE name is empty
        if table is None or str(table).strip() == "":
            continue

        rows.append({
            "db_name": current_db,
            "no": no,
            "table": str(table).strip(),
            "col_count": col_count,
            "row_count": row_count,
            "s2i_db": str(s2i_db).strip() if s2i_db else None,
            "s2i_table": str(s2i_table).strip() if s2i_table else None,
            "s2i_col_count": s2i_col_count,
            "s2i_row_count": s2i_row_count,
            "col_match": col_match,
            "row_match": row_match,
            "etc": str(etc).strip() if etc else None,
            "row_reason": str(row_reason).strip() if row_reason else None,
        })
    return rows


def read_all_excel_sheets(wb):
    """Read all 3 sheets and return {sheet_name: [rows]}."""
    result = {}
    for sn in SHEET_NAMES:
        ws = wb[sn]
        result[sn] = read_excel_sheet(ws)
    return result


# ──────────────────────────────────────────────
# 2. Read JSON files from raw/
# ──────────────────────────────────────────────

def read_json_files():
    """
    Read all JSON files in raw/ that represent DB schemas.
    Returns {db_name: {table_name: {"col_count": int, "row_count": int}}}
    """
    result = {}
    json_files = sorted(glob.glob(os.path.join(RAW_DIR, "*.json")))
    for jf in json_files:
        fname = os.path.basename(jf)
        # Skip non-schema JSON files
        if fname in ("data-samples.json", "ops-schema.json"):
            continue
        # Only include DB1_*, DB2_*, BROADCAST*, FALL_LEAGUE* pattern files
        if not (fname.startswith("DB1_") or fname.startswith("DB2_") or
                fname.startswith("BROADCAST") or fname.startswith("FALL_LEAGUE")):
            continue

        with open(jf, "r", encoding="utf-8") as f:
            data = json.load(f)

        db_name = data.get("db_name", fname.replace(".json", ""))
        tables_dict = data.get("tables", {})

        table_info = {}
        for tname, tdata in tables_dict.items():
            col_count = len(tdata.get("columns", []))
            row_count = tdata.get("row_count", 0)
            table_info[tname] = {"col_count": col_count, "row_count": row_count}

        result[db_name] = table_info
    return result


# ──────────────────────────────────────────────
# 3. Print helpers
# ──────────────────────────────────────────────

def fmt_val(v, width=10):
    """Format a value for table display."""
    s = str(v) if v is not None else "-"
    return s.rjust(width)


def fmt_match(v):
    """Format match status."""
    if v is True:
        return "  OK"
    elif v is False:
        return "DIFF"
    else:
        return "   -"


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

def main():
    print("=" * 120)
    print("  COMPARISON: Excel Table Inventory vs. Backup DB JSON Schemas")
    print("=" * 120)
    print()

    # ── Part 1: Excel ──
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    excel_data = read_all_excel_sheets(wb)

    all_excel_tables = set()  # (db_name, table_name) globally
    excel_by_db = {}          # db_name -> set of table names

    for sheet_name in SHEET_NAMES:
        rows = excel_data[sheet_name]
        print("-" * 120)
        print(f"  EXCEL SHEET: {sheet_name}  ({len(rows)} tables)")
        print("-" * 120)
        print(f"  {'DB Name':<35} {'No':>4}  {'Table':<40} {'BkCols':>6} {'BkRows':>10}  "
              f"{'S2iCols':>7} {'S2iRows':>10}  {'ColOK':>5} {'RowOK':>5}  {'ETC / Reason'}")
        print(f"  {'─'*35} {'─'*4}  {'─'*40} {'─'*6} {'─'*10}  "
              f"{'─'*7} {'─'*10}  {'─'*5} {'─'*5}  {'─'*30}")

        for r in rows:
            db = r["db_name"] or "(unknown)"
            no_str = str(r["no"]) if r["no"] is not None else "-"
            tbl = r["table"]
            bcols = r["col_count"] if r["col_count"] is not None else "-"
            brows = r["row_count"] if r["row_count"] is not None else "-"
            scols = r["s2i_col_count"] if r["s2i_col_count"] is not None else "-"
            srows = r["s2i_row_count"] if r["s2i_row_count"] is not None else "-"
            cm = fmt_match(r["col_match"])
            rm = fmt_match(r["row_match"])
            note_parts = []
            if r["etc"]:
                note_parts.append(r["etc"])
            if r["row_reason"]:
                note_parts.append(r["row_reason"])
            note = " | ".join(note_parts) if note_parts else ""

            print(f"  {db:<35} {no_str:>4}  {tbl:<40} {str(bcols):>6} {str(brows):>10}  "
                  f"{str(scols):>7} {str(srows):>10}  {cm:>5} {rm:>5}  {note}")

            all_excel_tables.add((db, tbl))
            excel_by_db.setdefault(db, set()).add(tbl)

        print()

    # Summary by DB from Excel
    print("=" * 80)
    print("  EXCEL SUMMARY: Table counts by DB name (all sheets combined)")
    print("=" * 80)
    for db in sorted(excel_by_db.keys()):
        print(f"    {db:<45} {len(excel_by_db[db]):>4} tables")
    print(f"\n    {'TOTAL UNIQUE (db+table) pairs:':<45} {len(all_excel_tables):>4}")
    # Also count unique table names ignoring DB
    unique_table_names_excel = set(t for _, t in all_excel_tables)
    print(f"    {'TOTAL UNIQUE table names (ignoring DB):':<45} {len(unique_table_names_excel):>4}")
    print()

    # ── Part 2: JSON backup DB ──
    json_data = read_json_files()

    print("=" * 120)
    print("  BACKUP DB (JSON files in raw/)")
    print("=" * 120)

    all_json_tables = set()  # (db_name, table_name)
    json_by_db = {}          # db_name -> set of table names

    for db_name in sorted(json_data.keys()):
        tables = json_data[db_name]
        json_by_db[db_name] = set(tables.keys())
        for tname in tables:
            all_json_tables.add((db_name, tname))

        print(f"\n  DB: {db_name}  ({len(tables)} tables)")
        print(f"    {'Table':<45} {'Cols':>6} {'Rows':>12}")
        print(f"    {'─'*45} {'─'*6} {'─'*12}")
        for tname in sorted(tables.keys()):
            info = tables[tname]
            print(f"    {tname:<45} {info['col_count']:>6} {info['row_count']:>12}")

    print()
    print("=" * 80)
    print("  BACKUP DB SUMMARY: Table counts by DB")
    print("=" * 80)
    for db in sorted(json_by_db.keys()):
        print(f"    {db:<45} {len(json_by_db[db]):>4} tables")
    print(f"\n    {'TOTAL UNIQUE (db+table) pairs:':<45} {len(all_json_tables):>4}")
    unique_table_names_json = set(t for _, t in all_json_tables)
    print(f"    {'TOTAL UNIQUE table names (ignoring DB):':<45} {len(unique_table_names_json):>4}")
    print()

    # ── Part 3: Cross-comparison ──
    print("=" * 120)
    print("  CROSS-COMPARISON: Excel vs. Backup DB (by DB name)")
    print("=" * 120)

    # Collect all DB names from both sources
    all_dbs = sorted(set(list(excel_by_db.keys()) + list(json_by_db.keys())))

    print(f"\n  {'DB Name':<45} {'Excel':>6} {'Backup':>7} {'Common':>7} {'ExcelOnly':>10} {'BackupOnly':>11}")
    print(f"  {'─'*45} {'─'*6} {'─'*7} {'─'*7} {'─'*10} {'─'*11}")

    total_excel_only = set()
    total_backup_only = set()
    total_common = set()

    for db in all_dbs:
        e_tables = excel_by_db.get(db, set())
        j_tables = json_by_db.get(db, set())

        # Case-insensitive matching: build lookup maps
        e_upper = {t.upper(): t for t in e_tables}
        j_upper = {t.upper(): t for t in j_tables}

        common_upper = set(e_upper.keys()) & set(j_upper.keys())
        excel_only_upper = set(e_upper.keys()) - set(j_upper.keys())
        backup_only_upper = set(j_upper.keys()) - set(e_upper.keys())

        common_count = len(common_upper)
        excel_only_count = len(excel_only_upper)
        backup_only_count = len(backup_only_upper)

        for t in common_upper:
            total_common.add((db, e_upper[t]))
        for t in excel_only_upper:
            total_excel_only.add((db, e_upper[t]))
        for t in backup_only_upper:
            total_backup_only.add((db, j_upper[t]))

        print(f"  {db:<45} {len(e_tables):>6} {len(j_tables):>7} {common_count:>7} {excel_only_count:>10} {backup_only_count:>11}")

    print()

    # Detail: Tables in Excel but NOT in backup DB
    print("=" * 120)
    print(f"  TABLES IN EXCEL BUT NOT IN BACKUP DB  ({len(total_excel_only)} tables)")
    print("=" * 120)
    if total_excel_only:
        for db, tbl in sorted(total_excel_only):
            print(f"    {db:<45} {tbl}")
    else:
        print("    (none)")
    print()

    # Detail: Tables in backup DB but NOT in Excel
    print("=" * 120)
    print(f"  TABLES IN BACKUP DB BUT NOT IN EXCEL  ({len(total_backup_only)} tables)")
    print("=" * 120)
    if total_backup_only:
        for db, tbl in sorted(total_backup_only):
            print(f"    {db:<45} {tbl}")
    else:
        print("    (none)")
    print()

    # ── Final totals ──
    print("=" * 80)
    print("  FINAL TOTALS")
    print("=" * 80)
    print(f"    Total unique (DB, table) pairs in Excel:       {len(all_excel_tables)}")
    print(f"    Total unique (DB, table) pairs in Backup DB:   {len(all_json_tables)}")
    print(f"    Total unique table names in Excel:             {len(unique_table_names_excel)}")
    print(f"    Total unique table names in Backup DB:         {len(unique_table_names_json)}")
    print(f"    Common (DB, table) pairs (case-insensitive):   {len(total_common)}")
    print(f"    In Excel only:                                 {len(total_excel_only)}")
    print(f"    In Backup DB only:                             {len(total_backup_only)}")
    print()


if __name__ == "__main__":
    main()
