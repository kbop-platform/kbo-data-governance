#!/usr/bin/env python3
"""
S2i_KBOP DB 보유 국제대회 목록 Excel 파일 읽기 스크립트

KBO 백업 DB에 보유 중인 국제대회 데이터 목록을 확인하기 위해
Excel 파일의 모든 시트, 헤더, 데이터를 출력한다.
"""

import openpyxl
import os
import sys


def read_international_excel(filepath: str) -> None:
    """Excel 파일의 모든 시트와 데이터를 출력한다."""

    if not os.path.exists(filepath):
        print(f"[ERROR] 파일을 찾을 수 없습니다: {filepath}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)

    print("=" * 80)
    print(f"파일: {os.path.basename(filepath)}")
    print(f"시트 수: {len(wb.sheetnames)}")
    print(f"시트 목록: {wb.sheetnames}")
    print("=" * 80)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print()
        print("-" * 80)
        print(f"시트: {sheet_name}")
        print(f"  행 수: {ws.max_row}, 열 수: {ws.max_column}")
        print("-" * 80)

        if ws.max_row is None or ws.max_row == 0:
            print("  (빈 시트)")
            continue

        # 모든 행 읽기
        all_rows = list(ws.iter_rows(values_only=True))

        if not all_rows:
            print("  (데이터 없음)")
            continue

        # 첫 번째 행을 헤더로 간주
        headers = all_rows[0]
        print()
        print(f"  [컬럼 헤더] ({len(headers)}개)")
        for i, header in enumerate(headers):
            print(f"    [{i}] {header}")

        print()
        print(f"  [데이터] (총 {len(all_rows) - 1}행)")
        print()

        # 각 데이터 행 출력
        for row_idx, row in enumerate(all_rows[1:], start=2):
            print(f"  --- 행 {row_idx} ---")
            for col_idx, (header, value) in enumerate(zip(headers, row)):
                print(f"    {header}: {value}")
            print()

        # 추가로 테이블 형태로도 출력
        print(f"  [테이블 뷰]")
        # 각 컬럼의 최대 너비 계산
        col_widths = []
        for col_idx in range(len(headers)):
            max_width = len(str(headers[col_idx] or ""))
            for row in all_rows[1:]:
                if col_idx < len(row):
                    max_width = max(max_width, len(str(row[col_idx] or "")))
            col_widths.append(min(max_width, 40))  # 최대 40자로 제한

        # 헤더 출력
        header_line = " | ".join(
            str(h or "").ljust(col_widths[i]) for i, h in enumerate(headers)
        )
        print(f"  {header_line}")
        print(f"  {' | '.join('-' * w for w in col_widths)}")

        # 데이터 행 출력
        for row in all_rows[1:]:
            row_line = " | ".join(
                str(row[i] if i < len(row) else "" or "").ljust(col_widths[i])
                for i in range(len(headers))
            )
            print(f"  {row_line}")

    wb.close()
    print()
    print("=" * 80)
    print("읽기 완료")
    print("=" * 80)


if __name__ == "__main__":
    filepath = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "S2i_KBOP DB 보유 국제대회 목록_260220.xlsx",
    )
    read_international_excel(filepath)
