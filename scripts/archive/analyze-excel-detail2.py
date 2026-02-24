#!/usr/bin/env python3
"""
Detailed drill-down analysis of S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx

PART A: 75 tables with column count differences (sorted by |diff|)
PART B: 56 backup-only tables with full details
PART C: 26 actual name differences (not just case)
PART D: 84 row count differences grouped by reason
PART E: S2i DB name mapping (backup DB -> S2i DB)
PART F: "전송 안함" (not transmitted) tables
"""

import openpyxl
import sys
from collections import defaultdict, Counter

EXCEL_PATH = "/home/user/dev/data-dict/S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx"

SEP = "=" * 120
SUB = "-" * 120


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def safe_int(val):
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def load_sheet_data(wb_formulas, wb_values, sheet_name):
    """
    Load and parse data from a single sheet.
    Uses wb_formulas for text values and wb_values for computed formula results.
    Forward-fills column A (backup DB name) for merged cells.
    Returns list of row dicts.
    """
    ws_f = wb_formulas[sheet_name]
    ws_v = wb_values[sheet_name]

    rows = []
    last_backup_db = ""

    for r in range(3, ws_f.max_row + 1):
        backup_db_raw = safe_str(ws_f.cell(row=r, column=1).value)
        no_val = ws_f.cell(row=r, column=2).value
        backup_table = safe_str(ws_f.cell(row=r, column=3).value)
        backup_cols = safe_int(ws_f.cell(row=r, column=4).value)
        backup_rows = safe_int(ws_f.cell(row=r, column=5).value)
        s2i_db = safe_str(ws_f.cell(row=r, column=6).value)
        s2i_table = safe_str(ws_f.cell(row=r, column=7).value)
        s2i_cols = safe_int(ws_f.cell(row=r, column=8).value)
        s2i_rows = safe_int(ws_f.cell(row=r, column=9).value)
        etc_notes = safe_str(ws_f.cell(row=r, column=12).value)
        row_reason = safe_str(ws_f.cell(row=r, column=13).value)

        # Computed values from data_only workbook
        col_match = ws_v.cell(row=r, column=10).value
        row_match = ws_v.cell(row=r, column=11).value

        # Forward-fill backup DB name (merged cells)
        if backup_db_raw:
            last_backup_db = backup_db_raw
        else:
            backup_db_raw = last_backup_db

        # Skip truly empty rows
        if not backup_table and not s2i_table:
            continue

        rows.append({
            "sheet": sheet_name,
            "row_num": r,
            "no": safe_int(no_val),
            "backup_db": backup_db_raw,
            "backup_table": backup_table,
            "backup_cols": backup_cols,
            "backup_rows": backup_rows,
            "s2i_db": s2i_db,
            "s2i_table": s2i_table,
            "s2i_cols": s2i_cols,
            "s2i_rows": s2i_rows,
            "col_match": col_match,
            "row_match": row_match,
            "etc_notes": etc_notes,
            "row_reason": row_reason,
        })

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# PART A: 75 tables with column count differences
# ─────────────────────────────────────────────────────────────────────────────
def part_a(all_rows):
    print(f"\n{SEP}")
    print("  PART A: TABLES WITH COLUMN COUNT DIFFERENCES")
    print(f"{SEP}")
    print("Tables where backup column count != S2i column count (both sides present).")
    print("Sorted by absolute difference (largest first).\n")

    comparable = [r for r in all_rows
                  if r["backup_cols"] is not None and r["s2i_cols"] is not None]
    diffs = [r for r in comparable if r["backup_cols"] != r["s2i_cols"]]

    # Sort by absolute diff descending, then by table name
    diffs.sort(key=lambda r: (-abs(r["s2i_cols"] - r["backup_cols"]),
                               r["backup_table"] or r["s2i_table"]))

    print(f"Total tables with column count differences: {len(diffs)}\n")

    # Print header
    hdr = (f"{'#':>3} | {'Sheet':<8} | {'Backup DB':<30} | {'Table':<35} | "
           f"{'Bkp Cols':>8} | {'S2i Cols':>8} | {'Diff':>6} | {'ETC Note'}")
    print(hdr)
    print("-" * len(hdr) + "-" * 30)

    for i, row in enumerate(diffs, 1):
        diff = row["s2i_cols"] - row["backup_cols"]
        diff_str = f"+{diff}" if diff > 0 else str(diff)
        table_name = row["backup_table"] or row["s2i_table"]
        etc = row["etc_notes"] if row["etc_notes"] else ""
        print(f"{i:>3} | {row['sheet']:<8} | {row['backup_db']:<30} | {table_name:<35} | "
              f"{row['backup_cols']:>8} | {row['s2i_cols']:>8} | {diff_str:>6} | {etc}")

    # Summary
    print(f"\n  --- Summary ---")
    s2i_more = [r for r in diffs if r["s2i_cols"] > r["backup_cols"]]
    bkp_more = [r for r in diffs if r["backup_cols"] > r["s2i_cols"]]
    print(f"  S2i has MORE columns than backup: {len(s2i_more)} tables")
    print(f"  Backup has MORE columns than S2i: {len(bkp_more)} tables")

    diff_vals = [abs(r["s2i_cols"] - r["backup_cols"]) for r in diffs]
    if diff_vals:
        print(f"  Max absolute difference: {max(diff_vals)}")
        print(f"  Min absolute difference: {min(diff_vals)}")
        print(f"  Avg absolute difference: {sum(diff_vals)/len(diff_vals):.1f}")

    # Distribution
    print(f"\n  Difference distribution:")
    dist = Counter(r["s2i_cols"] - r["backup_cols"] for r in diffs)
    for d, cnt in sorted(dist.items()):
        d_str = f"+{d}" if d > 0 else str(d)
        print(f"    {d_str:>4}: {cnt} table(s)")

    return diffs


# ─────────────────────────────────────────────────────────────────────────────
# PART B: 56 backup-only tables
# ─────────────────────────────────────────────────────────────────────────────
def part_b(all_rows):
    print(f"\n{SEP}")
    print("  PART B: BACKUP-ONLY TABLES (no S2i match)")
    print(f"{SEP}")
    print("Tables in backup DB (cols C/D/E populated) but NOT in S2i (cols F/G empty).\n")

    backup_only = [r for r in all_rows if r["backup_table"] and not r["s2i_table"]]

    print(f"Total backup-only tables: {len(backup_only)}\n")

    hdr = (f"{'#':>3} | {'Sheet':<8} | {'Backup DB':<30} | {'Backup Table':<40} | "
           f"{'Bkp Cols':>8} | {'Bkp Rows':>10} | {'ETC Note'}")
    print(hdr)
    print("-" * len(hdr) + "-" * 30)

    for i, row in enumerate(backup_only, 1):
        cols = row["backup_cols"] if row["backup_cols"] is not None else ""
        rows_val = row["backup_rows"] if row["backup_rows"] is not None else ""
        etc = row["etc_notes"] if row["etc_notes"] else ""
        print(f"{i:>3} | {row['sheet']:<8} | {row['backup_db']:<30} | {row['backup_table']:<40} | "
              f"{str(cols):>8} | {str(rows_val):>10} | {etc}")

    # Group by backup DB
    print(f"\n  --- Grouped by Backup DB ---")
    by_db = defaultdict(list)
    for r in backup_only:
        by_db[r["backup_db"]].append(r)
    for db in sorted(by_db.keys()):
        tables = by_db[db]
        print(f"\n  {db} ({len(tables)} tables):")
        for r in tables:
            cols = r["backup_cols"] if r["backup_cols"] is not None else "?"
            rows_val = r["backup_rows"] if r["backup_rows"] is not None else "?"
            etc = f" -- {r['etc_notes']}" if r["etc_notes"] else ""
            print(f"    - {r['backup_table']} (cols={cols}, rows={rows_val}){etc}")

    # Group by sheet
    print(f"\n  --- By Sheet ---")
    for sheet in ["DB1_", "DB2_", "전송 외"]:
        s_rows = [r for r in backup_only if r["sheet"] == sheet]
        print(f"  {sheet}: {len(s_rows)} backup-only tables")

    return backup_only


# ─────────────────────────────────────────────────────────────────────────────
# PART C: 26 actual name differences (not just case)
# ─────────────────────────────────────────────────────────────────────────────
def part_c(all_rows):
    print(f"\n{SEP}")
    print("  PART C: ACTUAL TABLE NAME DIFFERENCES (not just case)")
    print(f"{SEP}")
    print("Tables where backup table name != S2i table name, even case-insensitively.\n")

    both_sides = [r for r in all_rows if r["backup_table"] and r["s2i_table"]]

    # Actual name differences (not just case)
    actual_diffs = [r for r in both_sides
                    if r["backup_table"].upper() != r["s2i_table"].upper()]

    # Also compute case-only for reference
    case_only = [r for r in both_sides
                 if r["backup_table"] != r["s2i_table"]
                 and r["backup_table"].upper() == r["s2i_table"].upper()]

    print(f"Total actual name differences (case-insensitive mismatch): {len(actual_diffs)}")
    print(f"(For reference: case-only differences: {len(case_only)})\n")

    hdr = (f"{'#':>3} | {'Sheet':<8} | {'Backup DB':<25} | {'Backup Table':<35} | "
           f"{'S2i DB':<35} | {'S2i Table':<35} | {'Bkp Cols':>8} | {'S2i Cols':>8}")
    print(hdr)
    print("-" * len(hdr) + "-" * 10)

    for i, row in enumerate(actual_diffs, 1):
        bc = row["backup_cols"] if row["backup_cols"] is not None else ""
        sc = row["s2i_cols"] if row["s2i_cols"] is not None else ""
        print(f"{i:>3} | {row['sheet']:<8} | {row['backup_db']:<25} | {row['backup_table']:<35} | "
              f"{row['s2i_db']:<35} | {row['s2i_table']:<35} | {str(bc):>8} | {str(sc):>8}")

    # Analyze the nature of differences
    print(f"\n  --- Nature of Differences ---")
    for i, row in enumerate(actual_diffs, 1):
        bt = row["backup_table"]
        st = row["s2i_table"]
        # Try to categorize
        if bt.upper().replace("_", "") == st.upper().replace("_", ""):
            nature = "underscore difference"
        elif bt.upper() in st.upper() or st.upper() in bt.upper():
            nature = "prefix/suffix difference"
        else:
            nature = "completely different name"
        print(f"  {i:>3}. {bt} <-> {st}  [{nature}]")

    return actual_diffs


# ─────────────────────────────────────────────────────────────────────────────
# PART D: 84 row count differences grouped by reason
# ─────────────────────────────────────────────────────────────────────────────
def part_d(all_rows):
    print(f"\n{SEP}")
    print("  PART D: ROW COUNT DIFFERENCES GROUPED BY REASON")
    print(f"{SEP}")
    print("Tables where backup row count != S2i row count, grouped by reason (col 13).\n")

    comparable = [r for r in all_rows
                  if r["backup_rows"] is not None and r["s2i_rows"] is not None]
    diffs = [r for r in comparable if r["backup_rows"] != r["s2i_rows"]]

    print(f"Total tables with row count differences: {len(diffs)}\n")

    # Group by reason
    by_reason = defaultdict(list)
    for r in diffs:
        reason = r["row_reason"] if r["row_reason"] else "(no reason given)"
        by_reason[reason].append(r)

    # Also group by etc_notes for rows without reason
    for reason in sorted(by_reason.keys(), key=lambda x: (-len(by_reason[x]), x)):
        tables = by_reason[reason]
        print(f'{SUB}')
        print(f'Reason: "{reason}" ({len(tables)} tables)')
        print(f'{SUB}')
        for r in sorted(tables, key=lambda x: -abs(x["s2i_rows"] - x["backup_rows"])):
            diff = r["s2i_rows"] - r["backup_rows"]
            diff_str = f"+{diff}" if diff > 0 else str(diff)
            table_name = r["backup_table"] or r["s2i_table"]
            etc = f" [ETC: {r['etc_notes']}]" if r["etc_notes"] else ""
            print(f"  - {r['backup_db']}.{table_name} "
                  f"(backup: {r['backup_rows']:,} rows, s2i: {r['s2i_rows']:,} rows, diff: {diff_str})"
                  f"{etc}")
        print()

    # Overall direction
    print(f"\n  --- Direction Summary ---")
    s2i_more = [r for r in diffs if r["s2i_rows"] > r["backup_rows"]]
    bkp_more = [r for r in diffs if r["backup_rows"] > r["s2i_rows"]]
    print(f"  S2i has MORE rows:   {len(s2i_more)} tables")
    print(f"  Backup has MORE rows: {len(bkp_more)} tables")

    # Largest differences
    print(f"\n  --- Top 10 Largest Row Differences ---")
    top10 = sorted(diffs, key=lambda r: -abs(r["s2i_rows"] - r["backup_rows"]))[:10]
    for i, r in enumerate(top10, 1):
        diff = r["s2i_rows"] - r["backup_rows"]
        diff_str = f"+{diff:,}" if diff > 0 else f"{diff:,}"
        table_name = r["backup_table"] or r["s2i_table"]
        reason = r["row_reason"] if r["row_reason"] else "(no reason)"
        print(f"  {i:>3}. {r['backup_db']}.{table_name}: "
              f"backup={r['backup_rows']:,}, s2i={r['s2i_rows']:,}, diff={diff_str} -- {reason}")

    return diffs


# ─────────────────────────────────────────────────────────────────────────────
# PART E: S2i DB name mapping
# ─────────────────────────────────────────────────────────────────────────────
def part_e(all_rows):
    print(f"\n{SEP}")
    print("  PART E: S2i DB NAME MAPPING")
    print(f"{SEP}")
    print("For each Backup DB name, shows the corresponding S2i DB name(s) and table counts.\n")

    # Build mapping: backup_db -> { s2i_db -> set of common tables }
    mapping = defaultdict(lambda: defaultdict(set))
    backup_only_tables = defaultdict(set)

    for r in all_rows:
        if r["backup_db"] and r["s2i_db"] and r["backup_table"] and r["s2i_table"]:
            table_name = r["backup_table"]
            mapping[r["backup_db"]][r["s2i_db"]].add(table_name)
        elif r["backup_db"] and r["backup_table"] and not r["s2i_table"]:
            backup_only_tables[r["backup_db"]].add(r["backup_table"])

    # Print table
    hdr = f"{'#':>3} | {'Backup DB':<35} | {'S2i DB':<45} | {'Tables in Common':>16}"
    print(hdr)
    print("-" * len(hdr) + "-" * 10)

    idx = 0
    for bdb in sorted(mapping.keys()):
        for sdb in sorted(mapping[bdb].keys()):
            idx += 1
            count = len(mapping[bdb][sdb])
            print(f"{idx:>3} | {bdb:<35} | {sdb:<45} | {count:>16}")

    # Detail: which tables map where
    print(f"\n  --- Detailed DB Mapping ---\n")
    for bdb in sorted(mapping.keys()):
        total_tables_in_bdb = sum(1 for r in all_rows
                                  if r["backup_db"] == bdb and r["backup_table"])
        bonly = len(backup_only_tables.get(bdb, set()))
        print(f"  Backup DB: {bdb} (total {total_tables_in_bdb} tables, {bonly} backup-only)")
        for sdb in sorted(mapping[bdb].keys()):
            tables = sorted(mapping[bdb][sdb])
            print(f"    -> S2i DB: {sdb} ({len(tables)} tables)")
            for t in tables:
                print(f"         {t}")
        if bdb in backup_only_tables:
            print(f"    -> (backup-only, no S2i match: {bonly} tables)")
            for t in sorted(backup_only_tables[bdb]):
                print(f"         {t}")
        print()

    return mapping


# ─────────────────────────────────────────────────────────────────────────────
# PART F: "전송 안함" (not transmitted) tables
# ─────────────────────────────────────────────────────────────────────────────
def part_f(all_rows):
    print(f"\n{SEP}")
    print('  PART F: "전송 안함" (NOT TRANSMITTED) TABLES')
    print(f"{SEP}")
    print('Tables marked as "전송 안함" in the ETC notes column.\n')

    # Search for "전송 안함" in etc_notes (also check row_reason)
    not_transmitted = []
    for r in all_rows:
        if "전송 안함" in r["etc_notes"] or "전송 안함" in r["row_reason"]:
            not_transmitted.append(r)

    # Also check for variations: "전송안함" (without space)
    for r in all_rows:
        if r not in not_transmitted:
            if "전송안함" in r["etc_notes"] or "전송안함" in r["row_reason"]:
                not_transmitted.append(r)

    # Also look for the sheet named "전송 외" which might contain these
    jeonsong_sheet = [r for r in all_rows if r["sheet"] == "전송 외"]

    print(f'Tables with "전송 안함" in notes: {len(not_transmitted)}')
    print(f'Tables in "전송 외" sheet: {len(jeonsong_sheet)}\n')

    if not_transmitted:
        print(f'  --- Tables marked "전송 안함" ---\n')
        hdr = (f"{'#':>3} | {'Sheet':<8} | {'Backup DB':<30} | {'Backup Table':<40} | "
               f"{'Bkp Cols':>8} | {'Bkp Rows':>10} | {'S2i DB':<35} | {'S2i Table':<35} | "
               f"{'S2i Cols':>8} | {'S2i Rows':>10} | {'ETC Note':<25} | {'Row Reason'}")
        print(hdr)
        print("-" * len(hdr) + "-" * 10)

        for i, r in enumerate(not_transmitted, 1):
            bc = r["backup_cols"] if r["backup_cols"] is not None else ""
            br = r["backup_rows"] if r["backup_rows"] is not None else ""
            sc = r["s2i_cols"] if r["s2i_cols"] is not None else ""
            sr = r["s2i_rows"] if r["s2i_rows"] is not None else ""
            etc = r["etc_notes"] if r["etc_notes"] else ""
            reason = r["row_reason"] if r["row_reason"] else ""
            sdb = r["s2i_db"] if r["s2i_db"] else ""
            st = r["s2i_table"] if r["s2i_table"] else ""
            print(f"{i:>3} | {r['sheet']:<8} | {r['backup_db']:<30} | {r['backup_table']:<40} | "
                  f"{str(bc):>8} | {str(br):>10} | {sdb:<35} | {st:<35} | "
                  f"{str(sc):>8} | {str(sr):>10} | {etc:<25} | {reason}")
    else:
        print('  No tables found with "전송 안함" in notes.')

    # Also show ALL entries in "전송 외" sheet for context
    if jeonsong_sheet:
        print(f'\n  --- All entries in "전송 외" sheet (for context) ---\n')
        hdr2 = (f"{'#':>3} | {'Backup DB':<30} | {'Backup Table':<40} | "
                f"{'Bkp Cols':>8} | {'Bkp Rows':>10} | {'S2i DB':<35} | {'S2i Table':<35} | "
                f"{'S2i Cols':>8} | {'S2i Rows':>10} | {'ETC Note':<25} | {'Row Reason'}")
        print(hdr2)
        print("-" * len(hdr2) + "-" * 10)

        for i, r in enumerate(jeonsong_sheet, 1):
            bc = r["backup_cols"] if r["backup_cols"] is not None else ""
            br = r["backup_rows"] if r["backup_rows"] is not None else ""
            sc = r["s2i_cols"] if r["s2i_cols"] is not None else ""
            sr = r["s2i_rows"] if r["s2i_rows"] is not None else ""
            etc = r["etc_notes"] if r["etc_notes"] else ""
            reason = r["row_reason"] if r["row_reason"] else ""
            sdb = r["s2i_db"] if r["s2i_db"] else ""
            st = r["s2i_table"] if r["s2i_table"] else ""
            print(f"{i:>3} | {r['backup_db']:<30} | {r['backup_table']:<40} | "
                  f"{str(bc):>8} | {str(br):>10} | {sdb:<35} | {st:<35} | "
                  f"{str(sc):>8} | {str(sr):>10} | {etc:<25} | {reason}")

    return not_transmitted


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print(SEP)
    print("  DETAILED DRILL-DOWN ANALYSIS")
    print(f"  File: {EXCEL_PATH}")
    print(f"  Date: 2026-02-23")
    print(SEP)

    print("\nLoading Excel workbook...")
    wb_formulas = openpyxl.load_workbook(EXCEL_PATH)
    wb_values = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    sheet_names = wb_formulas.sheetnames
    print(f"Sheets found: {sheet_names}")

    all_rows = []
    for sname in sheet_names:
        sheet_rows = load_sheet_data(wb_formulas, wb_values, sname)
        print(f"  Sheet '{sname}': {len(sheet_rows)} data rows loaded")
        all_rows.extend(sheet_rows)

    print(f"\nTotal rows loaded: {len(all_rows)}")

    # Quick sanity check counts
    both = [r for r in all_rows if r["backup_table"] and r["s2i_table"]]
    bonly = [r for r in all_rows if r["backup_table"] and not r["s2i_table"]]
    sonly = [r for r in all_rows if not r["backup_table"] and r["s2i_table"]]
    print(f"  Both sides present: {len(both)}")
    print(f"  Backup-only: {len(bonly)}")
    print(f"  S2i-only: {len(sonly)}")

    # Run all parts
    part_a(all_rows)
    part_b(all_rows)
    part_c(all_rows)
    part_d(all_rows)
    part_e(all_rows)
    part_f(all_rows)

    print(f"\n{SEP}")
    print("  ANALYSIS COMPLETE")
    print(SEP)


if __name__ == "__main__":
    main()
