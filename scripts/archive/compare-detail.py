#!/usr/bin/env python3
"""
compare-detail.py

S2i_KBOP 백업 DB 테이블별 데이터 확인 엑셀 파일과
OPS 스키마 JSON (ops-schema.json)을 비교 분석한다.

SECTION 1: Excel 백업 테이블 고유 이름 (col 2)
SECTION 2: Excel Sports2i 테이블 고유 이름 (col 6)
SECTION 3: OPS 테이블 중 Excel에 없는 것
SECTION 4: Excel 테이블 중 OPS에 없는 것
SECTION 5: 이름이 매칭되는 테이블
"""

import json
import os
from collections import defaultdict

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx")
OPS_JSON_PATH = os.path.join(BASE_DIR, "raw", "ops-schema.json")

# ── 1. Read OPS JSON ──────────────────────────────────────────────────────
with open(OPS_JSON_PATH, "r", encoding="utf-8") as f:
    ops_tables = json.load(f)

# Build lookup: upper-cased physical_name -> table info
ops_by_upper = {}
for t in ops_tables:
    real_cols = [c for c in t["columns"] if c["physical_name"] != "NN"]
    ops_by_upper[t["physical_name"].upper()] = {
        "physical_name": t["physical_name"],
        "logical_name": t["logical_name"],
        "category": t["category"] or "(없음)",
        "col_count": len(real_cols),
    }

# ── 2. Read Excel ─────────────────────────────────────────────────────────
SHEET_NAMES = ["DB1_", "DB2_", "전송 외"]
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# Each row becomes a dict
excel_rows = []

for sn in SHEET_NAMES:
    ws = wb[sn]
    last_backup_db = None
    last_s2i_db = None

    for r in range(3, ws.max_row + 1):
        vals = []
        for c in range(1, 14):
            vals.append(ws.cell(r, c).value)

        # Forward-fill col 0 (backup DB) and col 5 (S2i DB)
        if vals[0] is not None:
            last_backup_db = str(vals[0]).strip()
        if vals[5] is not None:
            last_s2i_db = str(vals[5]).strip()

        backup_table = str(vals[2]).strip() if vals[2] is not None else ""
        s2i_table = str(vals[6]).strip() if vals[6] is not None else ""

        # Skip rows where both table names are empty
        if not backup_table and not s2i_table:
            continue

        row_dict = {
            "sheet": sn,
            "backup_db": last_backup_db or "",
            "backup_table": backup_table,
            "backup_cols": vals[3],
            "backup_rows": vals[4],
            "s2i_db": last_s2i_db or "",
            "s2i_table": s2i_table,
            "s2i_cols": vals[7],
            "s2i_rows": vals[8],
            "col_match": vals[9],
            "row_match": vals[10],
            "etc_note": str(vals[11]).strip() if vals[11] is not None else "",
            "row_reason": str(vals[12]).strip() if vals[12] is not None else "",
        }
        excel_rows.append(row_dict)

wb.close()

# ── 3. Build helper sets ──────────────────────────────────────────────────
# Backup table names: upper -> list of (original_name, sheet, db)
backup_names = defaultdict(list)  # upper -> [(name, sheet, db), ...]
for row in excel_rows:
    if row["backup_table"]:
        key = row["backup_table"].upper()
        backup_names[key].append((row["backup_table"], row["sheet"], row["backup_db"]))

# S2i table names: upper -> list of (original_name, sheet, db, corresponding_backup)
s2i_names = defaultdict(list)
for row in excel_rows:
    if row["s2i_table"]:
        key = row["s2i_table"].upper()
        s2i_names[key].append((row["s2i_table"], row["sheet"], row["s2i_db"], row["backup_table"]))

# Combined Excel names (upper) for matching
all_excel_upper = set(backup_names.keys()) | set(s2i_names.keys())

# ══════════════════════════════════════════════════════════════════════════
# SECTION 1: Excel unique backup table names (col 2)
# ══════════════════════════════════════════════════════════════════════════
print("=" * 90)
print("SECTION 1: Excel 고유 백업 테이블 이름 (col 2)")
print("=" * 90)

sorted_backup = sorted(backup_names.keys())
print(f"\n총 {len(sorted_backup)}개 고유 이름 (대소문자 무시)\n")
print(f"{'#':<4} {'테이블명':<40} {'시트':<12} {'DB'}")
print("-" * 90)
for i, key in enumerate(sorted_backup, 1):
    entries = backup_names[key]
    # Deduplicate (sheet, db) pairs
    seen = set()
    locations = []
    for name, sheet, db in entries:
        loc = (sheet, db)
        if loc not in seen:
            seen.add(loc)
            locations.append(loc)
    display_name = entries[0][0]  # Use first original casing
    for j, (sheet, db) in enumerate(locations):
        if j == 0:
            print(f"{i:<4} {display_name:<40} {sheet:<12} {db}")
        else:
            print(f"{'':>4} {'':>40} {sheet:<12} {db}")

# ══════════════════════════════════════════════════════════════════════════
# SECTION 2: Excel unique Sports2i table names (col 6)
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 90)
print("SECTION 2: Excel 고유 Sports2i 테이블 이름 (col 6)")
print("=" * 90)

sorted_s2i = sorted(s2i_names.keys())
print(f"\n총 {len(sorted_s2i)}개 고유 이름 (대소문자 무시)\n")
print(f"{'#':<4} {'S2i 테이블명':<40} {'대응 백업 테이블':<30} {'시트':<12} {'DB'}")
print("-" * 120)
for i, key in enumerate(sorted_s2i, 1):
    entries = s2i_names[key]
    # Deduplicate
    seen = set()
    locations = []
    for name, sheet, db, bk_table in entries:
        loc = (sheet, db, bk_table)
        if loc not in seen:
            seen.add(loc)
            locations.append((name, sheet, db, bk_table))
    display_name = entries[0][0]
    for j, (name, sheet, db, bk_table) in enumerate(locations):
        if j == 0:
            print(f"{i:<4} {display_name:<40} {bk_table:<30} {sheet:<12} {db}")
        else:
            print(f"{'':>4} {'':>40} {bk_table:<30} {sheet:<12} {db}")

# ══════════════════════════════════════════════════════════════════════════
# SECTION 3: OPS tables NOT in Excel (neither col2 nor col6)
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 90)
print("SECTION 3: OPS 테이블 중 Excel에 없는 것 (col2, col6 모두 미존재)")
print("=" * 90)

ops_not_in_excel = {}
for upper_name, info in ops_by_upper.items():
    if upper_name not in all_excel_upper:
        ops_not_in_excel[upper_name] = info

# Group by category
by_category = defaultdict(list)
for upper_name, info in ops_not_in_excel.items():
    by_category[info["category"]].append(info)

# Sort categories, then tables within each
total_missing = sum(len(v) for v in by_category.values())
print(f"\n총 {total_missing}개 OPS 테이블이 Excel에 없음\n")

for cat in sorted(by_category.keys()):
    tables = sorted(by_category[cat], key=lambda x: x["physical_name"])
    print(f"\n  [{cat}] ({len(tables)}개)")
    print(f"  {'#':<4} {'테이블명':<45} {'논리명':<35} {'컬럼수':>6}")
    print("  " + "-" * 92)
    for j, info in enumerate(tables, 1):
        print(f"  {j:<4} {info['physical_name']:<45} {info['logical_name']:<35} {info['col_count']:>6}")

# ══════════════════════════════════════════════════════════════════════════
# SECTION 4: Excel tables NOT in OPS (neither backup nor S2i names)
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 90)
print("SECTION 4: Excel 테이블 중 OPS에 없는 것")
print("=" * 90)

ops_upper_set = set(ops_by_upper.keys())

# Collect Excel names not in OPS
excel_not_in_ops = {}

# Check backup names
for upper_name in backup_names:
    if upper_name not in ops_upper_set:
        if upper_name not in excel_not_in_ops:
            excel_not_in_ops[upper_name] = {"sources": []}
        for name, sheet, db in backup_names[upper_name]:
            excel_not_in_ops[upper_name]["sources"].append(
                {"name": name, "sheet": sheet, "db": db, "side": "백업"}
            )

# Check S2i names
for upper_name in s2i_names:
    if upper_name not in ops_upper_set:
        if upper_name not in excel_not_in_ops:
            excel_not_in_ops[upper_name] = {"sources": []}
        for name, sheet, db, bk in s2i_names[upper_name]:
            excel_not_in_ops[upper_name]["sources"].append(
                {"name": name, "sheet": sheet, "db": db, "side": "S2i"}
            )

print(f"\n총 {len(excel_not_in_ops)}개 Excel 테이블이 OPS에 없음\n")
print(f"{'#':<4} {'테이블명':<40} {'출처':<6} {'시트':<12} {'DB'}")
print("-" * 100)

for i, upper_name in enumerate(sorted(excel_not_in_ops.keys()), 1):
    entry = excel_not_in_ops[upper_name]
    # Deduplicate sources
    seen = set()
    unique_sources = []
    for s in entry["sources"]:
        key = (s["name"], s["sheet"], s["db"], s["side"])
        if key not in seen:
            seen.add(key)
            unique_sources.append(s)
    for j, s in enumerate(unique_sources):
        if j == 0:
            print(f"{i:<4} {s['name']:<40} {s['side']:<6} {s['sheet']:<12} {s['db']}")
        else:
            print(f"{'':>4} {s['name']:<40} {s['side']:<6} {s['sheet']:<12} {s['db']}")

# ══════════════════════════════════════════════════════════════════════════
# SECTION 5: Name-matched tables (exact match, case-insensitive)
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 90)
print("SECTION 5: 이름 매칭 테이블 (OPS ↔ Excel, 대소문자 무시)")
print("=" * 90)

matched = {}
for upper_name in ops_upper_set:
    if upper_name in all_excel_upper:
        matched[upper_name] = ops_by_upper[upper_name]

print(f"\n총 {len(matched)}개 테이블 매칭\n")
print(f"{'#':<4} {'테이블명':<40} {'OPS 논리명':<30} {'OPS 컬럼':>8}  {'Excel 컬럼 (백업)':>18}  {'Excel 컬럼 (S2i)':>16}")
print("-" * 120)

for i, upper_name in enumerate(sorted(matched.keys()), 1):
    ops_info = matched[upper_name]

    # Get Excel column counts for this table
    backup_col_vals = set()
    s2i_col_vals = set()
    for row in excel_rows:
        if row["backup_table"].upper() == upper_name:
            if row["backup_cols"] is not None:
                backup_col_vals.add(row["backup_cols"])
        if row["s2i_table"].upper() == upper_name:
            if row["s2i_cols"] is not None:
                s2i_col_vals.add(row["s2i_cols"])

    bk_str = ", ".join(str(v) for v in sorted(backup_col_vals)) if backup_col_vals else "-"
    s2i_str = ", ".join(str(v) for v in sorted(s2i_col_vals)) if s2i_col_vals else "-"

    print(f"{i:<4} {ops_info['physical_name']:<40} {ops_info['logical_name']:<30} {ops_info['col_count']:>8}  {bk_str:>18}  {s2i_str:>16}")

# ── Summary ───────────────────────────────────────────────────────────────
print("\n" + "=" * 90)
print("요약")
print("=" * 90)
print(f"  OPS 테이블 수:                    {len(ops_by_upper)}")
print(f"  Excel 고유 백업 테이블 (col2):    {len(backup_names)}")
print(f"  Excel 고유 S2i 테이블 (col6):     {len(s2i_names)}")
print(f"  Excel 전체 고유 테이블:            {len(all_excel_upper)}")
print(f"  OPS에만 있는 테이블:              {total_missing}")
print(f"  Excel에만 있는 테이블:            {len(excel_not_in_ops)}")
print(f"  이름 매칭 테이블:                 {len(matched)}")
