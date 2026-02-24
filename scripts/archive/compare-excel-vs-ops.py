#!/usr/bin/env python3
"""
compare-excel-vs-ops.py

엑셀 파일(S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx)과
OPS PDF JSON(raw/ops-schema.json)의 테이블 목록을 비교한다.

비교 대상:
  A) 엑셀 백업 테이블(col 2) vs OPS
  B) 엑셀 Sports2i 테이블(col 6) vs OPS
  C) 엑셀 전체(col 2 + col 6 합집합) vs OPS
"""

import json
import os
from collections import defaultdict

import openpyxl

# ──────────────────────────────────────────────
# 경로 설정
# ──────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx")
OPS_JSON_PATH = os.path.join(BASE_DIR, "raw", "ops-schema.json")

HEADER_ROWS = 2  # rows 1-2 are headers


def load_excel():
    """
    엑셀에서 모든 시트를 읽어 백업 테이블(col2)과 Sports2i 테이블(col6) 목록을 추출한다.
    merged cell은 forward-fill 처리.

    Returns:
        backup_tables: list of dict  (table, cols, db, sheet)
        s2i_tables:    list of dict  (table, cols, db, sheet)
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    backup_tables = []
    s2i_tables = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        prev_db_backup = None   # forward-fill for col 0 (column A = index 1)
        prev_db_s2i = None      # forward-fill for col 5 (column F = index 6)

        for row_idx in range(HEADER_ROWS + 1, ws.max_row + 1):
            # Column indices (1-based in openpyxl)
            # col 0 -> column 1 (A) : backup DB name
            # col 2 -> column 3 (C) : backup TABLE name
            # col 3 -> column 4 (D) : backup COLUMN count
            # col 5 -> column 6 (F) : Sports2i DB name
            # col 6 -> column 7 (G) : Sports2i TABLE name
            # col 7 -> column 8 (H) : Sports2i COLUMN count

            c0 = ws.cell(row=row_idx, column=1).value   # backup DB
            c2 = ws.cell(row=row_idx, column=3).value   # backup TABLE
            c3 = ws.cell(row=row_idx, column=4).value   # backup COLUMN count
            c5 = ws.cell(row=row_idx, column=6).value   # s2i DB
            c6 = ws.cell(row=row_idx, column=7).value   # s2i TABLE
            c7 = ws.cell(row=row_idx, column=8).value   # s2i COLUMN count

            # Forward-fill DB names (merged cells)
            if c0 is not None:
                prev_db_backup = str(c0).strip()
            if c5 is not None:
                prev_db_s2i = str(c5).strip()

            # Backup table (col 2)
            if c2 is not None:
                tbl_name = str(c2).strip()
                if tbl_name:
                    cols_count = None
                    if c3 is not None:
                        try:
                            cols_count = int(c3)
                        except (ValueError, TypeError):
                            pass
                    backup_tables.append({
                        "table": tbl_name,
                        "cols": cols_count,
                        "db": prev_db_backup or "",
                        "sheet": sheet_name,
                    })

            # Sports2i table (col 6)
            if c6 is not None:
                tbl_name = str(c6).strip()
                if tbl_name:
                    cols_count = None
                    if c7 is not None:
                        try:
                            cols_count = int(c7)
                        except (ValueError, TypeError):
                            pass
                    s2i_tables.append({
                        "table": tbl_name,
                        "cols": cols_count,
                        "db": prev_db_s2i or "",
                        "sheet": sheet_name,
                    })

    return backup_tables, s2i_tables


def load_ops_json():
    """OPS JSON에서 테이블 목록을 로드한다."""
    with open(OPS_JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    ops_tables = []
    for entry in data:
        pname = entry.get("physical_name", "").strip()
        lname = entry.get("logical_name", "").strip()
        cat = entry.get("category", "").strip()
        columns = entry.get("columns", [])
        if pname:
            ops_tables.append({
                "table": pname,
                "logical_name": lname,
                "category": cat,
                "col_count": len(columns),
            })
    return ops_tables


def deduplicate_case_insensitive(table_list):
    """
    Case-insensitive 중복 제거. 원본 대소문자를 유지하되
    같은 이름(case-insensitive)이 여러 번 나오면 첫 번째 것을 유지.
    Returns: dict {upper_name: original_name}
    """
    seen = {}
    for t in table_list:
        upper = t["table"].upper()
        if upper not in seen:
            seen[upper] = t["table"]
    return seen


def build_detail_map(table_list, key_field="table"):
    """
    table_list 를 upper(table) -> [entries] 맵으로 변환.
    같은 이름이 여러 시트/DB에 걸쳐 등장할 수 있으므로 리스트로 저장.
    """
    m = defaultdict(list)
    for t in table_list:
        upper = t[key_field].upper()
        m[upper].append(t)
    return m


def compare(source_name, source_upper_set, source_detail, ops_upper_set, ops_detail):
    """3-way 비교를 수행하고 결과를 반환한다."""
    matched = source_upper_set & ops_upper_set
    only_source = source_upper_set - ops_upper_set
    only_ops = ops_upper_set - source_upper_set
    return {
        "source_name": source_name,
        "matched": sorted(matched),
        "only_source": sorted(only_source),
        "only_ops": sorted(only_ops),
        "source_detail": source_detail,
        "ops_detail": ops_detail,
    }


def print_separator(char="=", width=100):
    print(char * width)


def print_section(title):
    print()
    print_separator("=")
    print(f"  {title}")
    print_separator("=")


def print_subsection(title):
    print()
    print_separator("-", 80)
    print(f"  {title}")
    print_separator("-", 80)


def print_table_list(label, names_upper, detail_map, show_ops_info=False, ops_detail=None):
    """테이블 목록을 상세 정보와 함께 출력한다."""
    print(f"\n  {label}: {len(names_upper)}개")
    if not names_upper:
        print("    (없음)")
        return

    for i, name_upper in enumerate(sorted(names_upper), 1):
        if show_ops_info and ops_detail:
            entries = ops_detail.get(name_upper, [])
            if entries:
                e = entries[0]
                print(f"    {i:3d}. {e['table']:<40s} [category={e['category']}, cols={e['col_count']}]")
            else:
                print(f"    {i:3d}. {name_upper}")
        else:
            entries = detail_map.get(name_upper, [])
            if entries:
                # 같은 테이블이 여러 시트/DB에 걸쳐 있을 수 있음
                first = entries[0]
                dbs = sorted(set(e["db"] for e in entries if e["db"]))
                sheets = sorted(set(e["sheet"] for e in entries))
                cols_vals = sorted(set(e["cols"] for e in entries if e["cols"] is not None))
                cols_str = "/".join(str(c) for c in cols_vals) if cols_vals else "?"
                db_str = " | ".join(dbs) if dbs else "?"
                sheet_str = ", ".join(sheets)
                print(f"    {i:3d}. {first['table']:<40s} [sheet={sheet_str}, db={db_str}, cols={cols_str}]")
            else:
                print(f"    {i:3d}. {name_upper}")


def print_comparison(comp):
    """비교 결과를 출력한다."""
    src = comp["source_name"]
    matched = comp["matched"]
    only_source = comp["only_source"]
    only_ops = comp["only_ops"]
    source_detail = comp["source_detail"]
    ops_detail = comp["ops_detail"]

    print(f"\n  총 비교 결과:")
    print(f"    - 양쪽 모두 존재 (matched): {len(matched)}개")
    print(f"    - {src}에만 존재:           {len(only_source)}개")
    print(f"    - OPS에만 존재:             {len(only_ops)}개")

    # Matched tables
    print_subsection(f"양쪽 모두 존재하는 테이블 ({len(matched)}개)")
    if matched:
        print(f"    {'No':>4s}  {'테이블명(Excel)':<35s} {'테이블명(OPS)':<35s} {'Excel cols':>10s} {'OPS cols':>10s} {'일치?':>6s}")
        print(f"    {'----':>4s}  {'-----------------------------------':<35s} {'-----------------------------------':<35s} {'----------':>10s} {'----------':>10s} {'------':>6s}")
        for i, name_upper in enumerate(matched, 1):
            # Source side
            src_entries = source_detail.get(name_upper, [])
            if src_entries:
                src_name = src_entries[0]["table"]
                src_cols_vals = sorted(set(e["cols"] for e in src_entries if e["cols"] is not None))
                src_cols_str = "/".join(str(c) for c in src_cols_vals) if src_cols_vals else "?"
            else:
                src_name = name_upper
                src_cols_str = "?"

            # OPS side
            ops_entries = ops_detail.get(name_upper, [])
            if ops_entries:
                ops_name = ops_entries[0]["table"]
                ops_cols = ops_entries[0]["col_count"]
            else:
                ops_name = name_upper
                ops_cols = "?"

            # Compare column counts
            match_flag = ""
            if src_cols_vals and ops_cols != "?":
                if any(c == ops_cols for c in src_cols_vals):
                    match_flag = "O"
                else:
                    match_flag = "X"

            print(f"    {i:4d}  {src_name:<35s} {ops_name:<35s} {src_cols_str:>10s} {str(ops_cols):>10s} {match_flag:>6s}")

    # Source-only tables
    print_subsection(f"{src}에만 존재하는 테이블 ({len(only_source)}개)")
    print_table_list(f"{src} only", only_source, source_detail)

    # OPS-only tables
    print_subsection(f"OPS에만 존재하는 테이블 ({len(only_ops)}개)")
    print_table_list("OPS only", only_ops, ops_detail, show_ops_info=True, ops_detail=ops_detail)


def main():
    print_section("엑셀 vs OPS JSON 테이블 비교 스크립트")
    print(f"  엑셀 파일: {EXCEL_PATH}")
    print(f"  OPS JSON:  {OPS_JSON_PATH}")

    # ──────────────────────────────────────────────
    # 1. 데이터 로드
    # ──────────────────────────────────────────────
    backup_tables, s2i_tables = load_excel()
    ops_tables = load_ops_json()

    # ──────────────────────────────────────────────
    # 2. 엑셀 데이터 출력
    # ──────────────────────────────────────────────
    print_section("1. 엑셀 데이터 요약")

    # Backup tables (col 2)
    backup_unique = deduplicate_case_insensitive(backup_tables)
    backup_detail = build_detail_map(backup_tables)

    print_subsection(f"엑셀 백업 테이블 (col 2) - 전체 행: {len(backup_tables)}개, 유니크(case-insensitive): {len(backup_unique)}개")
    for i, (upper, orig) in enumerate(sorted(backup_unique.items()), 1):
        entries = backup_detail[upper]
        dbs = sorted(set(e["db"] for e in entries if e["db"]))
        sheets = sorted(set(e["sheet"] for e in entries))
        cols_vals = sorted(set(e["cols"] for e in entries if e["cols"] is not None))
        cols_str = "/".join(str(c) for c in cols_vals) if cols_vals else "?"
        db_str = " | ".join(dbs) if dbs else "?"
        print(f"    {i:3d}. {orig:<40s} [sheets={','.join(sheets)}, dbs={db_str}, cols={cols_str}]")

    # Sports2i tables (col 6)
    s2i_unique = deduplicate_case_insensitive(s2i_tables)
    s2i_detail = build_detail_map(s2i_tables)

    print_subsection(f"엑셀 Sports2i 테이블 (col 6) - 전체 행: {len(s2i_tables)}개, 유니크(case-insensitive): {len(s2i_unique)}개")
    for i, (upper, orig) in enumerate(sorted(s2i_unique.items()), 1):
        entries = s2i_detail[upper]
        dbs = sorted(set(e["db"] for e in entries if e["db"]))
        sheets = sorted(set(e["sheet"] for e in entries))
        cols_vals = sorted(set(e["cols"] for e in entries if e["cols"] is not None))
        cols_str = "/".join(str(c) for c in cols_vals) if cols_vals else "?"
        db_str = " | ".join(dbs) if dbs else "?"
        print(f"    {i:3d}. {orig:<40s} [sheets={','.join(sheets)}, dbs={db_str}, cols={cols_str}]")

    # ──────────────────────────────────────────────
    # 3. OPS JSON 데이터 출력
    # ──────────────────────────────────────────────
    ops_detail = build_detail_map(ops_tables)

    print_section(f"2. OPS JSON 테이블 목록 ({len(ops_tables)}개)")
    for i, t in enumerate(ops_tables, 1):
        print(f"    {i:3d}. {t['table']:<40s} [{t['category']:<15s}] cols={t['col_count']}")

    # ──────────────────────────────────────────────
    # 4. 비교
    # ──────────────────────────────────────────────
    backup_upper_set = set(backup_unique.keys())
    s2i_upper_set = set(s2i_unique.keys())
    ops_upper_set = set(t["table"].upper() for t in ops_tables)

    # Union of all Excel tables
    all_excel_tables = backup_tables + s2i_tables
    all_excel_unique = deduplicate_case_insensitive(all_excel_tables)
    all_excel_detail = build_detail_map(all_excel_tables)
    all_excel_upper_set = set(all_excel_unique.keys())

    # Comparison A: backup (col 2) vs OPS
    print_section("3. Comparison A: 엑셀 백업 테이블 (col 2) vs OPS")
    comp_a = compare("Excel백업(col2)", backup_upper_set, backup_detail, ops_upper_set, ops_detail)
    print_comparison(comp_a)

    # Comparison B: S2i (col 6) vs OPS
    print_section("4. Comparison B: 엑셀 Sports2i 테이블 (col 6) vs OPS")
    comp_b = compare("ExcelS2i(col6)", s2i_upper_set, s2i_detail, ops_upper_set, ops_detail)
    print_comparison(comp_b)

    # Comparison C: All Excel vs OPS
    print_section("5. Comparison C: 엑셀 전체 (col 2 + col 6 합집합) vs OPS")
    comp_c = compare("Excel전체(col2+col6)", all_excel_upper_set, all_excel_detail, ops_upper_set, ops_detail)
    print_comparison(comp_c)

    # ──────────────────────────────────────────────
    # 5. 요약 테이블
    # ──────────────────────────────────────────────
    print_section("6. 최종 요약")
    print()
    header = f"  {'Source':<30s} | {'Unique Tables':>15s} | {'In Both':>10s} | {'Only in Source':>15s} | {'Only in OPS':>12s}"
    sep    = f"  {'-'*30}-+-{'-'*15}-+-{'-'*10}-+-{'-'*15}-+-{'-'*12}"
    print(header)
    print(sep)

    rows = [
        ("Excel 백업 (col 2)", len(backup_unique), len(comp_a["matched"]), len(comp_a["only_source"]), len(comp_a["only_ops"])),
        ("Excel Sports2i (col 6)", len(s2i_unique), len(comp_b["matched"]), len(comp_b["only_source"]), len(comp_b["only_ops"])),
        ("Excel 전체 (col2+col6)", len(all_excel_unique), len(comp_c["matched"]), len(comp_c["only_source"]), len(comp_c["only_ops"])),
        ("OPS JSON", len(ops_tables), "-", "-", "-"),
    ]
    for name, unique, both, only_src, only_ops_val in rows:
        print(f"  {name:<30s} | {str(unique):>15s} | {str(both):>10s} | {str(only_src):>15s} | {str(only_ops_val):>12s}")

    print()
    print_separator("=")
    print("  비교 완료.")
    print_separator("=")


if __name__ == "__main__":
    main()
