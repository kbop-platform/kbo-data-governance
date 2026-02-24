#!/usr/bin/env python3
"""
Deep analysis of S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx

Compares backup DB tables with Sports2i DB tables across 3 sheets (DB1_, DB2_, 전송 외).
Produces 8 sections of analysis covering table names, column counts, row counts,
unmatched tables, DB name mappings, notes, and an overall summary.
"""

import openpyxl
import sys
from collections import defaultdict, Counter

EXCEL_PATH = "/home/user/dev/data-dict/S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx"

SEPARATOR = "=" * 100
SUB_SEP = "-" * 100


def safe_str(val):
    """Convert a cell value to a stripped string, or empty string if None."""
    if val is None:
        return ""
    return str(val).strip()


def safe_int(val):
    """Convert a cell value to an integer, or None if not possible."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def safe_bool(val):
    """Convert a cell value to a boolean representation string."""
    if val is None:
        return "(empty)"
    if isinstance(val, bool):
        return str(val)
    s = str(val).strip().upper()
    if s in ("TRUE", "1"):
        return "True"
    if s in ("FALSE", "0"):
        return "False"
    # It might be a formula string like "=D35=H35"
    if s.startswith("="):
        return f"(formula: {val})"
    return str(val)


def load_sheet_data(wb_formulas, wb_values, sheet_name):
    """
    Load and parse data from a single sheet.
    Uses wb_formulas for text values (col A, B, C, F, G, L, M)
    and wb_values for computed formula results (col J, K).
    Forward-fills column A (backup DB name).
    Returns list of row dicts.
    """
    ws_f = wb_formulas[sheet_name]
    ws_v = wb_values[sheet_name]

    rows = []
    last_backup_db = ""

    for r in range(3, ws_f.max_row + 1):
        # Read all columns from formula workbook (for text)
        backup_db_raw = safe_str(ws_f.cell(row=r, column=1).value)
        no_val = ws_f.cell(row=r, column=2).value
        backup_table = safe_str(ws_f.cell(row=r, column=3).value)
        backup_cols = safe_int(ws_f.cell(row=r, column=4).value)
        backup_rows = safe_int(ws_f.cell(row=r, column=5).value)
        s2i_db = safe_str(ws_f.cell(row=r, column=6).value)
        s2i_table = safe_str(ws_f.cell(row=r, column=7).value)
        s2i_cols = safe_int(ws_f.cell(row=r, column=8).value)
        s2i_rows = safe_int(ws_f.cell(row=r, column=9).value)
        etc_notes = safe_str(ws_f.cell(row=r, column=12).value)
        row_reason = safe_str(ws_f.cell(row=r, column=13).value)

        # Read computed values from data_only workbook
        col_match = ws_v.cell(row=r, column=10).value
        row_match = ws_v.cell(row=r, column=11).value

        # Forward-fill backup DB name
        if backup_db_raw:
            last_backup_db = backup_db_raw
        else:
            backup_db_raw = last_backup_db

        # Skip truly empty rows (no table name at all on either side)
        if not backup_table and not s2i_table:
            continue

        rows.append({
            "sheet": sheet_name,
            "row_num": r,
            "no": safe_int(no_val),
            "backup_db": backup_db_raw,
            "backup_table": backup_table,
            "backup_cols": backup_cols,
            "backup_rows": backup_rows,
            "s2i_db": s2i_db,
            "s2i_table": s2i_table,
            "s2i_cols": s2i_cols,
            "s2i_rows": s2i_rows,
            "col_match": col_match,
            "row_match": row_match,
            "etc_notes": etc_notes,
            "row_reason": row_reason,
        })

    return rows


def print_section_header(num, title):
    print(f"\n{SEPARATOR}")
    print(f"  SECTION {num}: {title}")
    print(SEPARATOR)


def analyze_table_name_diffs(all_rows):
    """SECTION 1: Table Name Differences."""
    print_section_header(1, "TABLE NAME DIFFERENCES")
    print("Comparing backup table name (col 2) vs S2i table name (col 6), case-insensitive.\n")

    diffs = []
    for row in all_rows:
        bt = row["backup_table"]
        st = row["s2i_table"]
        # Only compare when both exist
        if bt and st:
            if bt.upper() != st.upper():
                diffs.append(row)

    if not diffs:
        print("  No table name differences found (all names match case-insensitively).\n")
        return diffs

    print(f"  Found {len(diffs)} rows with table name differences:\n")
    print(f"  {'Sheet':<10} {'Row':<5} {'Backup DB':<35} {'Backup Table':<35} {'S2i DB':<45} {'S2i Table':<35}")
    print(f"  {'-'*10} {'-'*5} {'-'*35} {'-'*35} {'-'*45} {'-'*35}")
    for row in diffs:
        print(f"  {row['sheet']:<10} {row['row_num']:<5} {row['backup_db']:<35} {row['backup_table']:<35} {row['s2i_db']:<45} {row['s2i_table']:<35}")

    # Also show case-only diffs vs actual name diffs
    case_only = [r for r in diffs if r["backup_table"].upper() == r["s2i_table"].upper()]
    real_diffs = [r for r in diffs if r["backup_table"].upper() != r["s2i_table"].upper()]

    print(f"\n  Summary:")
    print(f"    Case-insensitive name differences: {len(real_diffs)}")
    # But let's also show case-sensitive diffs
    case_sensitive_diffs = [r for r in all_rows if r["backup_table"] and r["s2i_table"]
                           and r["backup_table"] != r["s2i_table"]]
    print(f"    Case-sensitive name differences (for reference): {len(case_sensitive_diffs)}")

    if case_sensitive_diffs:
        # Show case-only differences separately
        case_only_list = [r for r in case_sensitive_diffs
                         if r["backup_table"].upper() == r["s2i_table"].upper()]
        if case_only_list:
            print(f"\n  Case-only differences ({len(case_only_list)} tables -- same name, different casing):")
            print(f"  {'Sheet':<10} {'Row':<5} {'Backup Table':<35} {'S2i Table':<35}")
            print(f"  {'-'*10} {'-'*5} {'-'*35} {'-'*35}")
            for row in case_only_list:
                print(f"  {row['sheet']:<10} {row['row_num']:<5} {row['backup_table']:<35} {row['s2i_table']:<35}")

    return diffs


def analyze_column_diffs(all_rows):
    """SECTION 2: Column Count Differences."""
    print_section_header(2, "COLUMN COUNT DIFFERENCES")
    print("Comparing backup column count (col 3) vs S2i column count (col 7).\n")

    comparable = [r for r in all_rows if r["backup_cols"] is not None and r["s2i_cols"] is not None]
    diffs = [r for r in comparable if r["backup_cols"] != r["s2i_cols"]]
    matches = [r for r in comparable if r["backup_cols"] == r["s2i_cols"]]

    print(f"  Comparable rows (both sides have column counts): {len(comparable)}")
    print(f"  Exact column count matches: {len(matches)}")
    print(f"  Column count differences: {len(diffs)}")

    if diffs:
        print(f"\n  {'Sheet':<10} {'Row':<5} {'Backup DB':<35} {'Table':<30} {'Bkp Cols':<10} {'S2i Cols':<10} {'Diff':<8} {'Col Match':<12}")
        print(f"  {'-'*10} {'-'*5} {'-'*35} {'-'*30} {'-'*10} {'-'*10} {'-'*8} {'-'*12}")
        for row in diffs:
            diff = row["s2i_cols"] - row["backup_cols"]
            diff_str = f"+{diff}" if diff > 0 else str(diff)
            table_name = row["backup_table"] or row["s2i_table"]
            col_match_str = safe_bool(row["col_match"])
            print(f"  {row['sheet']:<10} {row['row_num']:<5} {row['backup_db']:<35} {table_name:<30} {row['backup_cols']:<10} {row['s2i_cols']:<10} {diff_str:<8} {col_match_str:<12}")

        # Distribution of differences
        diff_values = [r["s2i_cols"] - r["backup_cols"] for r in diffs]
        print(f"\n  Difference distribution:")
        for d, cnt in sorted(Counter(diff_values).items()):
            d_str = f"+{d}" if d > 0 else str(d)
            print(f"    {d_str}: {cnt} table(s)")

    return diffs, matches


def analyze_row_diffs(all_rows):
    """SECTION 3: Row Count Differences."""
    print_section_header(3, "ROW COUNT DIFFERENCES")
    print("Comparing backup row count (col 4) vs S2i row count (col 8).\n")

    comparable = [r for r in all_rows if r["backup_rows"] is not None and r["s2i_rows"] is not None]
    diffs = [r for r in comparable if r["backup_rows"] != r["s2i_rows"]]
    matches = [r for r in comparable if r["backup_rows"] == r["s2i_rows"]]

    print(f"  Comparable rows (both sides have row counts): {len(comparable)}")
    print(f"  Exact row count matches: {len(matches)}")
    print(f"  Row count differences: {len(diffs)}")

    if diffs:
        print(f"\n  {'Sheet':<10} {'Row':<5} {'Backup DB':<30} {'Table':<28} {'Bkp Rows':<12} {'S2i Rows':<12} {'Diff':<12} {'Row Match':<10} {'Reason'}")
        print(f"  {'-'*10} {'-'*5} {'-'*30} {'-'*28} {'-'*12} {'-'*12} {'-'*12} {'-'*10} {'-'*40}")
        for row in diffs:
            diff = row["s2i_rows"] - row["backup_rows"]
            diff_str = f"+{diff}" if diff > 0 else str(diff)
            table_name = row["backup_table"] or row["s2i_table"]
            row_match_str = safe_bool(row["row_match"])
            reason = row["row_reason"] if row["row_reason"] else ""
            print(f"  {row['sheet']:<10} {row['row_num']:<5} {row['backup_db']:<30} {table_name:<28} {row['backup_rows']:<12} {row['s2i_rows']:<12} {diff_str:<12} {row_match_str:<10} {reason}")

        # Categorize reasons
        reasons = [r["row_reason"] for r in diffs if r["row_reason"]]
        if reasons:
            print(f"\n  Reason categories ({len(reasons)} reasons provided out of {len(diffs)} differences):")
            reason_counts = Counter(reasons)
            for reason, cnt in reason_counts.most_common():
                print(f"    [{cnt:>3}x] {reason}")
        else:
            print(f"\n  No reasons provided for row count differences.")

        # Summary stats
        diffs_positive = [r for r in diffs if r["s2i_rows"] > r["backup_rows"]]
        diffs_negative = [r for r in diffs if r["s2i_rows"] < r["backup_rows"]]
        print(f"\n  Direction of differences:")
        print(f"    S2i has MORE rows than backup: {len(diffs_positive)}")
        print(f"    Backup has MORE rows than S2i: {len(diffs_negative)}")

    return diffs, matches


def analyze_s2i_only(all_rows):
    """SECTION 4: S2i-only Tables."""
    print_section_header(4, "S2i-ONLY TABLES")
    print("Tables that exist in S2i (col 6/7 populated) but NOT in backup (col 2/3 empty).\n")

    s2i_only = [r for r in all_rows if not r["backup_table"] and r["s2i_table"]]

    if not s2i_only:
        print("  No S2i-only tables found.\n")
        return s2i_only

    print(f"  Found {len(s2i_only)} S2i-only tables:\n")
    print(f"  {'Sheet':<10} {'Row':<5} {'S2i DB':<45} {'S2i Table':<35} {'S2i Cols':<10} {'S2i Rows':<12} {'Notes'}")
    print(f"  {'-'*10} {'-'*5} {'-'*45} {'-'*35} {'-'*10} {'-'*12} {'-'*30}")
    for row in s2i_only:
        cols = row["s2i_cols"] if row["s2i_cols"] is not None else ""
        rows_val = row["s2i_rows"] if row["s2i_rows"] is not None else ""
        notes = row["etc_notes"] if row["etc_notes"] else ""
        print(f"  {row['sheet']:<10} {row['row_num']:<5} {row['s2i_db']:<45} {row['s2i_table']:<35} {str(cols):<10} {str(rows_val):<12} {notes}")

    return s2i_only


def analyze_backup_only(all_rows):
    """SECTION 5: Backup-only Tables."""
    print_section_header(5, "BACKUP-ONLY TABLES")
    print("Tables in backup (col 2/3 populated) but NOT in S2i (col 6/7 empty).\n")

    backup_only = [r for r in all_rows if r["backup_table"] and not r["s2i_table"]]

    if not backup_only:
        print("  No backup-only tables found.\n")
        return backup_only

    print(f"  Found {len(backup_only)} backup-only tables:\n")
    print(f"  {'Sheet':<10} {'Row':<5} {'Backup DB':<35} {'Backup Table':<35} {'Bkp Cols':<10} {'Bkp Rows':<12} {'Notes'}")
    print(f"  {'-'*10} {'-'*5} {'-'*35} {'-'*35} {'-'*10} {'-'*12} {'-'*30}")
    for row in backup_only:
        cols = row["backup_cols"] if row["backup_cols"] is not None else ""
        rows_val = row["backup_rows"] if row["backup_rows"] is not None else ""
        notes = row["etc_notes"] if row["etc_notes"] else ""
        print(f"  {row['sheet']:<10} {row['row_num']:<5} {row['backup_db']:<35} {row['backup_table']:<35} {str(cols):<10} {str(rows_val):<12} {notes}")

    return backup_only


def analyze_db_names(all_rows):
    """SECTION 6: S2i DB Names."""
    print_section_header(6, "DB NAME ANALYSIS")
    print("Unique DB names on both backup and S2i sides.\n")

    # Backup DB names
    backup_dbs = defaultdict(int)
    for row in all_rows:
        if row["backup_db"]:
            backup_dbs[row["backup_db"]] += 1

    print(f"  BACKUP DB Names ({len(backup_dbs)} unique):")
    print(f"  {'DB Name':<45} {'Table Count'}")
    print(f"  {'-'*45} {'-'*12}")
    for db, cnt in sorted(backup_dbs.items(), key=lambda x: -x[1]):
        print(f"  {db:<45} {cnt}")

    # S2i DB names
    s2i_dbs = defaultdict(int)
    for row in all_rows:
        if row["s2i_db"]:
            s2i_dbs[row["s2i_db"]] += 1

    print(f"\n  S2i DB Names ({len(s2i_dbs)} unique):")
    print(f"  {'DB Name':<55} {'Table Count'}")
    print(f"  {'-'*55} {'-'*12}")
    for db, cnt in sorted(s2i_dbs.items(), key=lambda x: -x[1]):
        print(f"  {db:<55} {cnt}")

    # Mapping between backup and S2i DB names
    print(f"\n  Backup DB -> S2i DB Mapping:")
    db_mapping = defaultdict(set)
    for row in all_rows:
        if row["backup_db"] and row["s2i_db"]:
            db_mapping[row["backup_db"]].add(row["s2i_db"])

    for bdb in sorted(db_mapping.keys()):
        s2i_list = sorted(db_mapping[bdb])
        print(f"    {bdb}")
        for sdb in s2i_list:
            print(f"      -> {sdb}")

    # Backup DBs with NO S2i mapping
    no_s2i = set()
    for row in all_rows:
        if row["backup_db"] and not row["s2i_db"]:
            no_s2i.add(row["backup_db"])
    if no_s2i:
        print(f"\n  Backup DBs with tables that have NO S2i mapping:")
        for db in sorted(no_s2i):
            count = sum(1 for r in all_rows if r["backup_db"] == db and not r["s2i_db"])
            print(f"    {db}: {count} table(s)")

    return backup_dbs, s2i_dbs


def analyze_notes(all_rows):
    """SECTION 7: ETC Notes and ROW Reason Analysis."""
    print_section_header(7, "ETC NOTES & ROW REASON ANALYSIS")
    print("All non-empty values from col 11 (ETC notes) and col 12 (ROW reason).\n")

    # ETC notes
    etc_rows = [r for r in all_rows if r["etc_notes"]]
    print(f"  === ETC Notes (col 11) === ({len(etc_rows)} entries)")
    if etc_rows:
        # Group by note content
        etc_groups = defaultdict(list)
        for row in etc_rows:
            etc_groups[row["etc_notes"]].append(row)

        print(f"\n  Unique ETC note values ({len(etc_groups)}):\n")
        for note, rows in sorted(etc_groups.items(), key=lambda x: -len(x[1])):
            print(f"  [{len(rows):>3}x] \"{note}\"")
            for row in rows:
                table_name = row["backup_table"] or row["s2i_table"]
                print(f"         - Sheet: {row['sheet']}, Row: {row['row_num']}, "
                      f"Backup DB: {row['backup_db']}, Table: {table_name}")
    else:
        print("  (none)")

    # ROW reasons
    reason_rows = [r for r in all_rows if r["row_reason"]]
    print(f"\n  === ROW Reasons (col 12) === ({len(reason_rows)} entries)")
    if reason_rows:
        reason_groups = defaultdict(list)
        for row in reason_rows:
            reason_groups[row["row_reason"]].append(row)

        print(f"\n  Unique ROW reason values ({len(reason_groups)}):\n")
        for reason, rows in sorted(reason_groups.items(), key=lambda x: -len(x[1])):
            print(f"  [{len(rows):>3}x] \"{reason}\"")
            for row in rows:
                table_name = row["backup_table"] or row["s2i_table"]
                print(f"         - Sheet: {row['sheet']}, Row: {row['row_num']}, "
                      f"Backup DB: {row['backup_db']}, Table: {table_name}")
    else:
        print("  (none)")

    return etc_rows, reason_rows


def overall_summary(all_rows, name_diffs, col_diffs, col_matches, row_diffs, row_matches,
                    s2i_only, backup_only):
    """SECTION 8: Overall Summary."""
    print_section_header(8, "OVERALL SUMMARY")

    total = len(all_rows)
    both_sides = [r for r in all_rows if r["backup_table"] and r["s2i_table"]]

    # Tables with identical schema = same name (case-insensitive) AND same column count
    identical_schema = [
        r for r in both_sides
        if r["backup_table"].upper() == r["s2i_table"].upper()
        and r["backup_cols"] is not None
        and r["s2i_cols"] is not None
        and r["backup_cols"] == r["s2i_cols"]
    ]

    # Tables with same name AND same column count AND same row count
    fully_identical = [
        r for r in identical_schema
        if r["backup_rows"] is not None
        and r["s2i_rows"] is not None
        and r["backup_rows"] == r["s2i_rows"]
    ]

    # Per-sheet breakdown
    sheets = ["DB1_", "DB2_", "전송 외"]

    print(f"\n  GRAND TOTALS:")
    print(f"  {'-'*60}")
    print(f"  Total data rows across all sheets:           {total}")
    print(f"  Rows with both backup AND S2i tables:        {len(both_sides)}")
    print(f"  Backup-only tables (no S2i match):           {len(backup_only)}")
    print(f"  S2i-only tables (no backup):                 {len(s2i_only)}")
    print(f"  {'-'*60}")
    print(f"  Table name differences (case-insensitive):   {len(name_diffs)}")

    # Case-sensitive diffs for reference
    case_sensitive_name_diffs = [
        r for r in both_sides
        if r["backup_table"] != r["s2i_table"]
    ]
    print(f"  Table name differences (case-sensitive):     {len(case_sensitive_name_diffs)}")

    print(f"  Column count differences:                    {len(col_diffs)}")
    print(f"  Column count matches:                        {len(col_matches)}")
    print(f"  Row count differences:                       {len(row_diffs)}")
    print(f"  Row count matches:                           {len(row_matches)}")
    print(f"  {'-'*60}")
    print(f"  Identical schema (name + col count match):   {len(identical_schema)}")
    print(f"  Fully identical (name + cols + rows match):  {len(fully_identical)}")

    print(f"\n  PER-SHEET BREAKDOWN:")
    print(f"  {'Sheet':<12} {'Total':<8} {'Both':<8} {'BkpOnly':<10} {'S2iOnly':<10} {'NameDiff':<10} {'ColDiff':<10} {'RowDiff':<10}")
    print(f"  {'-'*12} {'-'*8} {'-'*8} {'-'*10} {'-'*10} {'-'*10} {'-'*10} {'-'*10}")
    for sheet in sheets:
        s_all = [r for r in all_rows if r["sheet"] == sheet]
        s_both = [r for r in both_sides if r["sheet"] == sheet]
        s_bkp_only = [r for r in backup_only if r["sheet"] == sheet]
        s_s2i_only = [r for r in s2i_only if r["sheet"] == sheet]
        s_name_diff = [r for r in name_diffs if r["sheet"] == sheet]
        s_col_diff = [r for r in col_diffs if r["sheet"] == sheet]
        s_row_diff = [r for r in row_diffs if r["sheet"] == sheet]
        print(f"  {sheet:<12} {len(s_all):<8} {len(s_both):<8} {len(s_bkp_only):<10} {len(s_s2i_only):<10} {len(s_name_diff):<10} {len(s_col_diff):<10} {len(s_row_diff):<10}")

    # Unique backup DBs
    backup_dbs = set(r["backup_db"] for r in all_rows if r["backup_db"])
    s2i_dbs = set(r["s2i_db"] for r in all_rows if r["s2i_db"])
    print(f"\n  Unique Backup DB names: {len(backup_dbs)}")
    print(f"  Unique S2i DB names:    {len(s2i_dbs)}")

    # Tables with data (non-zero row counts)
    backup_with_data = [r for r in all_rows if r["backup_rows"] is not None and r["backup_rows"] > 0]
    backup_empty = [r for r in all_rows if r["backup_rows"] is not None and r["backup_rows"] == 0]
    print(f"\n  Backup tables with data (rows > 0):  {len(backup_with_data)}")
    print(f"  Backup tables with NO data (rows=0): {len(backup_empty)}")
    if backup_empty:
        print(f"    Empty backup tables:")
        for r in backup_empty:
            print(f"      - [{r['sheet']}] {r['backup_db']}.{r['backup_table']}")


def main():
    print(SEPARATOR)
    print("  DEEP ANALYSIS: S2i_KBOP 백업 DB 테이블별 데이터 확인")
    print(f"  File: {EXCEL_PATH}")
    print(f"  Generated: 2026-02-23")
    print(SEPARATOR)

    # Load workbook twice: once for formulas (text), once for computed values
    print("\nLoading Excel workbook...")
    wb_formulas = openpyxl.load_workbook(EXCEL_PATH)
    wb_values = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    sheet_names = wb_formulas.sheetnames
    print(f"Sheets found: {sheet_names}")

    # Parse all sheets
    all_rows = []
    for sname in sheet_names:
        sheet_rows = load_sheet_data(wb_formulas, wb_values, sname)
        print(f"  Sheet '{sname}': {len(sheet_rows)} data rows loaded")
        all_rows.extend(sheet_rows)

    print(f"\nTotal rows loaded: {len(all_rows)}")

    # Run all 8 analysis sections
    name_diffs = analyze_table_name_diffs(all_rows)
    col_diffs, col_matches = analyze_column_diffs(all_rows)
    row_diffs, row_matches = analyze_row_diffs(all_rows)
    s2i_only = analyze_s2i_only(all_rows)
    backup_only = analyze_backup_only(all_rows)
    analyze_db_names(all_rows)
    analyze_notes(all_rows)
    overall_summary(all_rows, name_diffs, col_diffs, col_matches, row_diffs, row_matches,
                    s2i_only, backup_only)

    print(f"\n{SEPARATOR}")
    print("  ANALYSIS COMPLETE")
    print(SEPARATOR)


if __name__ == "__main__":
    main()
