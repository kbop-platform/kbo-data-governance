#!/usr/bin/env python3
"""
S2i_KBOP 백업 DB 테이블별 데이터 확인 엑셀 파일 분석 스크립트

엑셀 파일을 열어 모든 시트의 구조와 데이터를 출력한다.
- 시트별 행/열 수, 헤더, 샘플 데이터
- 테이블명이 포함된 컬럼 탐색 및 유니크 테이블명 카운트
"""

import openpyxl
import os
import sys

EXCEL_PATH = "/home/user/dev/data-dict/S2i_KBOP 백업 DB 테이블별 데이터 확인_260220.xlsx"


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: File not found: {EXCEL_PATH}")
        sys.exit(1)

    print(f"파일: {EXCEL_PATH}")
    print(f"파일 크기: {os.path.getsize(EXCEL_PATH):,} bytes")
    print("=" * 100)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    print(f"\n총 시트 수: {len(wb.sheetnames)}")
    print(f"시트 목록: {wb.sheetnames}")
    print("=" * 100)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'#' * 100}")
        print(f"## 시트: [{sheet_name}]")
        print(f"{'#' * 100}")
        print(f"  행(rows): {ws.max_row}, 열(columns): {ws.max_column}")
        print(f"  데이터 행 수 (헤더 제외): {max(0, ws.max_row - 1)}")

        # 전체 데이터를 리스트로 변환
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            print("  (빈 시트)")
            continue

        # 헤더 추출
        headers = all_rows[0]
        print(f"\n  ### 컬럼 헤더 ({len(headers)}개):")
        for i, h in enumerate(headers):
            print(f"    [{i}] {h}")

        # 첫 5행 데이터
        data_rows = all_rows[1:]
        sample_count = min(5, len(data_rows))
        print(f"\n  ### 첫 {sample_count}행 데이터:")
        for row_idx, row in enumerate(data_rows[:sample_count], start=1):
            print(f"  --- Row {row_idx} ---")
            for col_idx, (header, value) in enumerate(zip(headers, row)):
                print(f"    [{col_idx}] {header}: {value}")

        # 테이블명 관련 컬럼 탐색
        # 테이블명이 들어있을 법한 컬럼 헤더 키워드
        table_keywords = ["table", "테이블", "TABLE", "Table", "tbl", "TBL", "명", "name", "NAME"]
        db_keywords = ["db", "DB", "database", "DATABASE", "데이터베이스", "스키마", "schema"]

        print(f"\n  ### 테이블명/DB명 관련 컬럼 탐색:")
        table_columns_found = []
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_str = str(h)
            for kw in table_keywords:
                if kw in h_str:
                    table_columns_found.append((i, h_str))
                    break

        db_columns_found = []
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_str = str(h)
            for kw in db_keywords:
                if kw in h_str:
                    db_columns_found.append((i, h_str))
                    break

        if table_columns_found:
            print(f"  테이블명 관련 컬럼:")
            for col_idx, col_name in table_columns_found:
                unique_vals = set()
                for row in data_rows:
                    if col_idx < len(row) and row[col_idx] is not None:
                        unique_vals.add(str(row[col_idx]).strip())
                unique_vals.discard("")
                print(f"    컬럼 [{col_idx}] '{col_name}': 유니크 값 {len(unique_vals)}개")
                # 유니크 값이 적으면 전부, 많으면 처음 30개만
                if len(unique_vals) <= 50:
                    for v in sorted(unique_vals):
                        print(f"      - {v}")
                else:
                    for v in sorted(list(unique_vals))[:30]:
                        print(f"      - {v}")
                    print(f"      ... (외 {len(unique_vals) - 30}개)")
        else:
            print("  테이블명 관련 컬럼 없음")

        if db_columns_found:
            print(f"  DB명 관련 컬럼:")
            for col_idx, col_name in db_columns_found:
                unique_vals = set()
                for row in data_rows:
                    if col_idx < len(row) and row[col_idx] is not None:
                        unique_vals.add(str(row[col_idx]).strip())
                unique_vals.discard("")
                print(f"    컬럼 [{col_idx}] '{col_name}': 유니크 값 {len(unique_vals)}개")
                if len(unique_vals) <= 50:
                    for v in sorted(unique_vals):
                        print(f"      - {v}")
                else:
                    for v in sorted(list(unique_vals))[:30]:
                        print(f"      - {v}")
                    print(f"      ... (외 {len(unique_vals) - 30}개)")
        else:
            print("  DB명 관련 컬럼 없음")

        # 추가: 모든 컬럼의 유니크 값 수 및 값이 테이블명처럼 보이는 경우 탐색
        # (테이블명은 보통 대문자+밑줄 패턴: TB_, TBL_, dbo. 등)
        print(f"\n  ### 모든 컬럼 유니크 값 수:")
        for i, h in enumerate(headers):
            if h is None:
                continue
            unique_vals = set()
            for row in data_rows:
                if i < len(row) and row[i] is not None:
                    unique_vals.add(str(row[i]).strip())
            unique_vals.discard("")
            print(f"    [{i}] {h}: {len(unique_vals)}개 유니크 값")

            # 테이블명 패턴 감지 (TB_, TBL_, dbo. 등)
            table_like = [v for v in unique_vals if any(
                v.upper().startswith(prefix) for prefix in ["TB_", "TBL_", "TBL", "DBO.", "T_"]
            )]
            if table_like and len(table_like) >= 3:
                print(f"        *** 테이블명 패턴 감지! ({len(table_like)}개) ***")
                for v in sorted(table_like)[:20]:
                    print(f"          - {v}")
                if len(table_like) > 20:
                    print(f"          ... (외 {len(table_like) - 20}개)")

    # 마지막: 전체 행 데이터 덤프 (시트 행이 적은 경우)
    print("\n" + "=" * 100)
    print("## 전체 데이터 덤프 (행이 200개 이하인 시트)")
    print("=" * 100)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) <= 201:  # 헤더 + 200행 이하
            print(f"\n--- 시트: [{sheet_name}] (총 {len(all_rows)}행) ---")
            headers = all_rows[0] if all_rows else []
            data_rows = all_rows[1:]
            # 테이블 형태로 출력
            # 헤더
            header_str = " | ".join(str(h) if h is not None else "" for h in headers)
            print(f"  {header_str}")
            print(f"  {'-' * len(header_str)}")
            for row in data_rows:
                row_str = " | ".join(str(v) if v is not None else "" for v in row)
                print(f"  {row_str}")
        else:
            print(f"\n--- 시트: [{sheet_name}] ({len(all_rows)}행 - 생략) ---")

    wb.close()
    print("\n\n완료.")


if __name__ == "__main__":
    main()
